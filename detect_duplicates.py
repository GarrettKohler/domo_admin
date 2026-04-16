"""Detect duplicate datasets in Domo and add a 'Duplicate Analysis' tab to the workspace planner."""

import json
import re
import logging
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Styles (matching project conventions)
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="3E5170", end_color="3E5170", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)

# Recommendation color coding
REC_FILLS = {
    "Keep":             PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid"),
    "Review":           PatternFill(start_color="F9A825", end_color="F9A825", fill_type="solid"),
    "Flag for Deletion": PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid"),
}
REC_FONTS = {
    "Keep":             Font(name="Arial", size=10, color="FFFFFF"),
    "Review":           Font(name="Arial", size=10, color="000000"),
    "Flag for Deletion": Font(name="Arial", size=10, color="FFFFFF"),
}

DUP_TYPE_FILLS = {
    "exact":     PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid"),
    "near":      PatternFill(start_color="E65100", end_color="E65100", fill_type="solid"),
    "versioned": PatternFill(start_color="F9A825", end_color="F9A825", fill_type="solid"),
}
DUP_TYPE_FONTS = {
    "exact":     Font(name="Arial", size=10, color="FFFFFF"),
    "near":      Font(name="Arial", size=10, color="FFFFFF"),
    "versioned": Font(name="Arial", size=10, color="000000"),
}

# Group row separator fill
GROUP_FILL = PatternFill(start_color="E3E8F0", end_color="E3E8F0", fill_type="solid")

CACHE_PATH = Path(__file__).parent / ".cache" / "latest.json"
OUTPUT_PATH = Path(__file__).parent / "output" / "domo_workspace_plan_20260411.xlsx"

# ---------------------------------------------------------------------------
# Normalization helpers
# ---------------------------------------------------------------------------

# Patterns stripped for near-duplicate matching
VERSION_PATTERN = re.compile(
    r"""
    \s*[-_]?\s*                       # optional separator
    (?:
        [Vv]\s*\d+(?:\.\d+)*          # V2, v2.0, V2.1, v 3
        |copy\s*\d*                    # copy, copy2, Copy 3
        |(?:re)?build\s*\d*            # build2, rebuild
        |final\s*\d*                   # final, final2
        |new\s*\d*                     # new, new2
        |latest                        # latest
        |updated                       # updated
        |revised                       # revised
        |backup                        # backup
        |bak                           # bak
    )
    \s*$                               # at end of string
    """,
    re.IGNORECASE | re.VERBOSE,
)

OLD_SUFFIX_PATTERN = re.compile(
    r"""
    \s*[-_]?\s*                        # optional separator
    (?:
        old                            # old
        |archive[d]?                   # archive, archived
        |deprecated                    # deprecated
        |original                      # original
        |prev(?:ious)?                 # prev, previous
        |legacy                        # legacy
        |retired                       # retired
        |\[?DNU\]?                     # DNU, [DNU]
        |\[?flag\s*for\s*deletion\]?   # Flag for Deletion
    )
    \s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

TRAILING_NUMBER = re.compile(r"\s*[-_]?\s*\d+\s*$")


def normalize_name(name: str) -> str:
    """Normalize a dataset name for near-duplicate comparison.

    Strips version suffixes, trailing numbers, extra whitespace, and lowercases.
    """
    n = name.strip()
    # Remove old/archive suffixes
    n = OLD_SUFFIX_PATTERN.sub("", n)
    # Remove version suffixes
    n = VERSION_PATTERN.sub("", n)
    # Remove trailing numbers (but not if the entire name is a number)
    candidate = TRAILING_NUMBER.sub("", n)
    if candidate.strip():
        n = candidate
    # Normalize whitespace and case
    n = re.sub(r"\s+", " ", n).strip().lower()
    # Remove trailing separators
    n = re.sub(r"[-_\s]+$", "", n)
    return n


def _parse_date(date_str):
    """Parse a date string, return datetime or None."""
    if not date_str:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%d"):
        try:
            return datetime.strptime(date_str.rstrip("Z"), fmt.rstrip("Z"))
        except (ValueError, AttributeError):
            continue
    return None


def _safe_row_count(ds):
    """Return row_count as int, defaulting to 0."""
    rc = ds.get("row_count")
    if rc is None:
        return 0
    try:
        return int(rc)
    except (ValueError, TypeError):
        return 0


# ---------------------------------------------------------------------------
# Duplicate detection
# ---------------------------------------------------------------------------

def find_duplicates(datasets: list[dict]) -> list[dict]:
    """Find all duplicate groups across three categories.

    Returns a flat list of records with group_number and duplicate_type assigned.
    """
    results = []
    group_num = 0
    seen_ids = set()  # Track dataset IDs already assigned to a group

    # -----------------------------------------------------------------------
    # 1. Exact name duplicates
    # -----------------------------------------------------------------------
    name_groups: dict[str, list[dict]] = defaultdict(list)
    for ds in datasets:
        name_groups[ds.get("dataset_name", "")].append(ds)

    exact_groups = {name: dups for name, dups in name_groups.items() if len(dups) > 1}

    for name, dups in sorted(exact_groups.items(), key=lambda x: -len(x[1])):
        group_num += 1
        group_ids = {d["dataset_id"] for d in dups}
        seen_ids.update(group_ids)
        _assign_recommendations(dups, group_num, "exact", results)

    # -----------------------------------------------------------------------
    # 2. Near duplicates (after normalization, excluding exact dupes already found)
    # -----------------------------------------------------------------------
    norm_groups: dict[str, list[dict]] = defaultdict(list)
    for ds in datasets:
        norm = normalize_name(ds.get("dataset_name", ""))
        norm_groups[norm].append(ds)

    for norm_name, dups in sorted(norm_groups.items(), key=lambda x: -len(x[1])):
        if len(dups) < 2:
            continue
        # Check if all members were already captured as exact duplicates
        dup_ids = {d["dataset_id"] for d in dups}
        unseen = dup_ids - seen_ids
        if not unseen:
            continue
        # Check if all share the exact same name (already handled above)
        unique_names = {d.get("dataset_name", "") for d in dups}
        if len(unique_names) == 1:
            continue

        group_num += 1
        seen_ids.update(dup_ids)

        # Determine if this is near or versioned
        dup_type = _classify_group(dups)
        _assign_recommendations(dups, group_num, dup_type, results)

    return results


def _classify_group(dups: list[dict]) -> str:
    """Classify a normalized-name group as 'near' or 'versioned'."""
    names = [d.get("dataset_name", "") for d in dups]

    # Check for version indicators
    version_indicators = re.compile(
        r"(?i)(?:[Vv]\s*\d|copy|old|archive|deprecated|original|prev|legacy|retired|DNU|flag.?for.?deletion|backup|new\d|build\d|final\d)",
    )
    has_version = sum(1 for n in names if version_indicators.search(n))

    # Check for trailing numbers that differ
    trailing_nums = []
    for n in names:
        m = re.search(r"(\d+)\s*$", n.strip())
        if m:
            trailing_nums.append(int(m.group(1)))

    if has_version >= 1:
        return "versioned"
    if len(set(trailing_nums)) > 1 and len(trailing_nums) == len(names):
        return "versioned"

    return "near"


def _assign_recommendations(dups: list[dict], group_num: int, dup_type: str, results: list[dict]):
    """Score each dataset in a duplicate group and assign recommendations."""
    # Score each dataset: higher = more likely the "active" version
    scored = []
    for ds in dups:
        dt = _parse_date(ds.get("data_current_at"))
        rc = _safe_row_count(ds)
        # Score: prefer most recently updated, then highest row count
        date_score = dt.timestamp() if dt else 0
        scored.append((date_score, rc, ds))

    scored.sort(key=lambda x: (x[0], x[1]), reverse=True)

    # The top-scoring dataset is recommended to Keep
    # Others get Review or Flag for Deletion
    best_date_score = scored[0][0]
    best_row_count = scored[0][1]

    for rank, (date_score, rc, ds) in enumerate(scored):
        if rank == 0:
            rec = "Keep"
        elif dup_type == "exact":
            # Exact duplicates: the non-best ones are likely redundant
            if date_score == 0 and rc == 0:
                rec = "Flag for Deletion"
            elif date_score < best_date_score * 0.5:
                rec = "Flag for Deletion"
            else:
                rec = "Review"
        elif dup_type == "versioned":
            # Versioned: old versions flagged, similar ones reviewed
            name_lower = ds.get("dataset_name", "").lower()
            if any(tag in name_lower for tag in ["old", "archive", "deprecated", "dnu", "flag for deletion",
                                                   "legacy", "retired", "prev", "original", "backup"]):
                rec = "Flag for Deletion"
            elif re.search(r"[Vv]\s*1(?:\b|\.0)", ds.get("dataset_name", "")):
                rec = "Review"
            else:
                rec = "Review"
        else:
            # Near duplicates: recommend review
            rec = "Review"

        results.append({
            "group_number": group_num,
            "dataset_id": ds.get("dataset_id", ""),
            "dataset_name": ds.get("dataset_name", ""),
            "owner_name": ds.get("owner_name", ""),
            "row_count": _safe_row_count(ds),
            "data_current_at": ds.get("data_current_at", ""),
            "duplicate_type": dup_type,
            "recommendation": rec,
        })


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

COLUMNS = [
    "group_number", "dataset_id", "dataset_name", "owner_name",
    "row_count", "data_current_at", "duplicate_type", "recommendation",
]


def _apply_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"


def _auto_fit(ws, max_width=60):
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), 1):
        max_len = 0
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)


def write_duplicate_analysis(wb, records: list[dict]) -> None:
    """Add the 'Duplicate Analysis' sheet to an existing workbook."""
    # Remove existing sheet if present
    if "Duplicate Analysis" in wb.sheetnames:
        del wb["Duplicate Analysis"]

    ws = wb.create_sheet("Duplicate Analysis")

    # Header row
    for col_idx, col in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=col)

    dup_type_col = COLUMNS.index("duplicate_type") + 1
    rec_col = COLUMNS.index("recommendation") + 1
    group_col = COLUMNS.index("group_number") + 1

    prev_group = None
    for row_idx, rec in enumerate(records, 2):
        current_group = rec["group_number"]

        for col_idx, col in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=rec.get(col, ""))
            cell.font = BODY_FONT

        # Alternate group background for readability
        if current_group != prev_group and prev_group is not None and current_group % 2 == 0:
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = GROUP_FILL

        # Color duplicate_type
        dt = rec.get("duplicate_type", "")
        dt_cell = ws.cell(row=row_idx, column=dup_type_col)
        if dt in DUP_TYPE_FILLS:
            dt_cell.fill = DUP_TYPE_FILLS[dt]
            dt_cell.font = DUP_TYPE_FONTS.get(dt, BODY_FONT)

        # Color recommendation
        r = rec.get("recommendation", "")
        r_cell = ws.cell(row=row_idx, column=rec_col)
        if r in REC_FILLS:
            r_cell.fill = REC_FILLS[r]
            r_cell.font = REC_FONTS.get(r, BODY_FONT)

        prev_group = current_group

    _apply_header(ws)
    _auto_fit(ws)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Loading dataset cache...", end=" ", flush=True)
    with open(CACHE_PATH) as f:
        cache = json.load(f)
    datasets = cache["datasets"]
    print(f"{len(datasets)} datasets loaded.")

    print("Detecting duplicates...", end=" ", flush=True)
    records = find_duplicates(datasets)

    if not records:
        print("No duplicates found.")
        return

    # Count groups
    group_nums = {r["group_number"] for r in records}
    total_groups = len(group_nums)

    # Count by type
    type_counts = defaultdict(int)
    type_group_counts = defaultdict(set)
    for r in records:
        type_counts[r["duplicate_type"]] += 1
        type_group_counts[r["duplicate_type"]].add(r["group_number"])

    print(f"{total_groups} duplicate groups found ({len(records)} total datasets).")

    # Summary
    print("\n--- Duplicate Analysis Summary ---")
    print(f"Total duplicate groups: {total_groups}")
    print(f"Total datasets involved: {len(records)}")
    for dt in ["exact", "near", "versioned"]:
        gcount = len(type_group_counts.get(dt, set()))
        dcount = type_counts.get(dt, 0)
        if gcount:
            print(f"  {dt:10s}: {gcount:3d} groups ({dcount:3d} datasets)")

    # Recommendation breakdown
    rec_counts = defaultdict(int)
    for r in records:
        rec_counts[r["recommendation"]] += 1
    print("\nRecommendations:")
    for rec_label in ["Keep", "Review", "Flag for Deletion"]:
        print(f"  {rec_label:20s}: {rec_counts.get(rec_label, 0):3d} datasets")

    # Write to workbook
    print(f"\nOpening workbook: {OUTPUT_PATH} ...", end=" ", flush=True)
    wb = load_workbook(str(OUTPUT_PATH))
    print("done.")

    print("Writing 'Duplicate Analysis' tab...", end=" ", flush=True)
    write_duplicate_analysis(wb, records)
    wb.save(str(OUTPUT_PATH))
    print("done.")
    print(f"Saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
