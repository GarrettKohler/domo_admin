#!/usr/bin/env python3
"""Build a Schema Similarity & Consolidation workbook.

Takes the output of schema_similarity.py and organizes it into an actionable
Excel workbook with:
  - Executive Summary tab (counts, top opportunities)
  - Consolidation Groups tab (clusters of related datasets, not just pairs)
  - Full Pair Details tab (every similar pair with shared columns)
  - Domain Breakdown tab (similarity by domain)

This is a separate deliverable from the cleanup rollout — cleanup is about
removing stale/unused items; consolidation is about merging active duplicates.
"""

import csv
import json
import os
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

CACHE_PATH = Path(__file__).resolve().parent / ".cache" / "latest.json"
SIMILARITY_CSV = Path(__file__).resolve().parent / "output" / "schema_similarity_analysis.csv"
OUTPUT_PATH = Path(__file__).resolve().parent / "output" / "domo_consolidation_report_20260412.xlsx"

# Styles
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LIKELY_FILL = PatternFill("solid", fgColor="FFC7CE")       # Red — likely duplicate
PROBABLE_FILL = PatternFill("solid", fgColor="FFE0B2")      # Orange — probable
SIMILAR_FILL = PatternFill("solid", fgColor="FFF9C4")       # Yellow — similar
GROUP_HEADER_FILL = PatternFill("solid", fgColor="E3F2FD")  # Light blue — group header
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def load_pairs() -> list[dict]:
    """Load similarity pairs from CSV."""
    with open(SIMILARITY_CSV, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def build_groups(pairs: list[dict], min_overlap: float = 75.0) -> list[list[str]]:
    """Cluster datasets into consolidation groups using union-find.

    Two datasets are in the same group if they share >= min_overlap% schema.
    """
    # Filter to strong pairs
    strong = [p for p in pairs if float(p["overlap_pct"]) >= min_overlap]

    # Union-Find
    parent: dict[str, str] = {}

    def find(x):
        while parent.get(x, x) != x:
            parent[x] = parent.get(parent[x], parent[x])
            x = parent[x]
        return x

    def union(x, y):
        px, py = find(x), find(y)
        if px != py:
            parent[px] = py

    # Collect all dataset IDs
    all_ids = set()
    for p in strong:
        a_id = p["dataset_a_id"]
        b_id = p["dataset_b_id"]
        all_ids.add(a_id)
        all_ids.add(b_id)
        union(a_id, b_id)

    # Collect groups
    groups_map: dict[str, list[str]] = defaultdict(list)
    for ds_id in all_ids:
        root = find(ds_id)
        groups_map[root].append(ds_id)

    # Sort groups by size descending, filter singles
    groups = [sorted(g) for g in groups_map.values() if len(g) >= 2]
    groups.sort(key=lambda g: -len(g))

    return groups


def main():
    # Load data
    pairs = load_pairs()
    print(f"Loaded {len(pairs)} similar pairs")

    with open(CACHE_PATH) as f:
        cache = json.load(f)

    ds_lookup = {ds["dataset_id"]: ds for ds in cache["datasets"]}

    # Build groups
    groups = build_groups(pairs, min_overlap=75.0)
    print(f"Identified {len(groups)} consolidation groups "
          f"({sum(len(g) for g in groups)} total datasets)")

    # Create workbook
    wb = Workbook()

    # ── Executive Summary ──
    ws_sum = wb.active
    ws_sum.title = "Executive Summary"
    ws_sum.sheet_properties.tabColor = "1F4E79"

    likely = [p for p in pairs if "Likely" in p.get("recommendation", "")]
    probable = [p for p in pairs if "Probable" in p.get("recommendation", "")]
    similar = [p for p in pairs if "Similar" in p.get("recommendation", "")]

    # Domain breakdown for summary
    domain_counts = Counter(p["domain"] for p in pairs)

    summary_lines = [
        ("Schema Similarity & Consolidation Report", ""),
        ("", ""),
        ("Prepared by: Aaron Olson", ""),
        ("Date: April 12, 2026", ""),
        ("", ""),
        ("OVERVIEW", ""),
        ("Datasets analyzed", len(ds_lookup)),
        ("Similar pairs found", len(pairs)),
        ("  Likely Duplicates (>=90% overlap)", len(likely)),
        ("  Probable Duplicates (>=80% overlap)", len(probable)),
        ("  Similar Schema (>=65% overlap)", len(similar)),
        ("", ""),
        ("Consolidation groups identified", len(groups)),
        ("Datasets in consolidation groups", sum(len(g) for g in groups)),
        ("Largest group size", max(len(g) for g in groups) if groups else 0),
        ("", ""),
        ("HOW TO USE THIS REPORT", ""),
        ("1. Review the Consolidation Groups tab for clusters of similar datasets", ""),
        ("2. For each group, determine which dataset to keep (newest, most rows)", ""),
        ("3. Retire/redirect duplicates to the canonical dataset", ""),
        ("4. Check the Pair Details tab for shared column specifics", ""),
        ("", ""),
        ("IMPACT", ""),
        ("Consolidating likely duplicates could reduce datasets by up to:",
         f"{sum(len(g) - 1 for g in groups)} datasets"),
        ("", ""),
        ("TOP DOMAINS WITH DUPLICATES", ""),
    ]

    for dom, count in domain_counts.most_common(10):
        summary_lines.append((f"  {dom}", count))

    for row_idx, (label, value) in enumerate(summary_lines, 1):
        cell_a = ws_sum.cell(row=row_idx, column=1, value=label)
        cell_b = ws_sum.cell(row=row_idx, column=2, value=value if value != "" else None)

        if row_idx == 1:
            cell_a.font = Font(bold=True, size=16, color="1F4E79")
        elif label in ("OVERVIEW", "HOW TO USE THIS REPORT", "IMPACT", "TOP DOMAINS WITH DUPLICATES"):
            cell_a.font = Font(bold=True, size=12, color="1F4E79")
        elif label.startswith("  "):
            cell_a.font = Font(size=11, color="666666")
            cell_a.alignment = Alignment(indent=2)
        else:
            cell_a.font = Font(size=11)

    ws_sum.column_dimensions["A"].width = 55
    ws_sum.column_dimensions["B"].width = 20

    # ── Consolidation Groups ──
    ws_grp = wb.create_sheet("Consolidation Groups")
    ws_grp.sheet_properties.tabColor = "C00000"

    grp_headers = [
        "Group", "Dataset Name", "Owner", "Rows", "Columns",
        "Last Data Update", "Domain", "Staleness", "Keep Candidate",
    ]
    for col_idx, h in enumerate(grp_headers, 1):
        cell = ws_grp.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    row_num = 2
    for grp_idx, group in enumerate(groups, 1):
        # Get dataset info for each member
        members = []
        for ds_id in group:
            ds = ds_lookup.get(ds_id, {})
            rows_count = ds.get("row_count", 0) or 0
            current = ds.get("data_current_at", "")
            members.append({
                "ds_id": ds_id,
                "name": ds.get("dataset_name", ""),
                "owner": ds.get("owner_name", ""),
                "rows": rows_count,
                "columns": ds.get("column_count", 0) or 0,
                "last_update": current[:10] if current else "Never",
                "current_ts": current or "",
            })

        # Sort: most recent data first, then most rows
        members.sort(key=lambda m: (-len(m["current_ts"]), m["current_ts"]), reverse=False)
        members.sort(key=lambda m: m["current_ts"], reverse=True)

        # Best candidate = first after sort
        from analytics import _classify_domain
        domain, _ = _classify_domain(members[0]["name"]) if members else ("", "")

        for i, m in enumerate(members):
            is_keep = (i == 0)  # first = keep candidate
            data = [
                grp_idx if i == 0 else "",
                m["name"],
                m["owner"],
                m["rows"],
                m["columns"],
                m["last_update"],
                domain if i == 0 else "",
                "",
                "KEEP" if is_keep else "Review",
            ]
            for col_idx, val in enumerate(data, 1):
                cell = ws_grp.cell(row=row_num, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if i == 0:
                    cell.fill = GROUP_HEADER_FILL
                    if col_idx == 9:
                        cell.font = Font(bold=True, color="2E7D32")
                elif col_idx == 9:
                    cell.font = Font(color="C00000")
            row_num += 1

        # Add blank separator row between groups
        row_num += 1

    # Column widths
    grp_widths = [8, 50, 20, 12, 10, 15, 25, 14, 14]
    for i, w in enumerate(grp_widths, 1):
        ws_grp.column_dimensions[get_column_letter(i)].width = w
    ws_grp.freeze_panes = "A2"

    # ── Pair Details ──
    ws_pairs = wb.create_sheet("Pair Details")
    ws_pairs.sheet_properties.tabColor = "E65100"

    pair_headers = [
        "Dataset A", "Dataset B", "Overlap %", "Shared Columns",
        "Total Unique Columns", "Recommendation", "Domain",
        "A Owner", "A Rows", "A Last Update",
        "B Owner", "B Rows", "B Last Update",
        "Shared Column Names",
    ]
    for col_idx, h in enumerate(pair_headers, 1):
        cell = ws_pairs.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for row_idx, p in enumerate(pairs[:2000], 2):  # Cap at 2000 for file size
        rec = p.get("recommendation", "")
        row_data = [
            p["dataset_a_name"],
            p["dataset_b_name"],
            float(p["overlap_pct"]),
            int(p["shared_column_count"]),
            int(p["total_unique_columns"]),
            rec,
            p["domain"],
            p["dataset_a_owner"],
            int(p["dataset_a_rows"]) if p["dataset_a_rows"] else 0,
            p["dataset_a_last_update"][:10] if p["dataset_a_last_update"] else "",
            p["dataset_b_owner"],
            int(p["dataset_b_rows"]) if p["dataset_b_rows"] else 0,
            p["dataset_b_last_update"][:10] if p["dataset_b_last_update"] else "",
            p.get("shared_columns", ""),
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws_pairs.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            if "Likely" in rec:
                if col_idx <= 6:
                    cell.fill = LIKELY_FILL
            elif "Probable" in rec:
                if col_idx <= 6:
                    cell.fill = PROBABLE_FILL
            elif "Similar" in rec:
                if col_idx <= 6:
                    cell.fill = SIMILAR_FILL

    pair_widths = [45, 45, 10, 10, 10, 30, 25, 18, 12, 13, 18, 12, 13, 60]
    for i, w in enumerate(pair_widths, 1):
        ws_pairs.column_dimensions[get_column_letter(i)].width = w
    ws_pairs.freeze_panes = "A2"
    ws_pairs.auto_filter.ref = f"A1:{get_column_letter(len(pair_headers))}{min(len(pairs)+1, 2001)}"

    # ── Domain Breakdown ──
    ws_dom = wb.create_sheet("Domain Breakdown")
    ws_dom.sheet_properties.tabColor = "2E7D32"

    dom_headers = ["Domain", "Total Pairs", "Likely Duplicates", "Probable", "Similar", "Datasets in Groups"]
    for col_idx, h in enumerate(dom_headers, 1):
        cell = ws_dom.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    # Compute per-domain stats
    domain_stats: dict[str, dict] = defaultdict(lambda: {
        "total": 0, "likely": 0, "probable": 0, "similar": 0, "ds_ids": set()
    })
    for p in pairs:
        dom = p["domain"]
        domain_stats[dom]["total"] += 1
        if "Likely" in p.get("recommendation", ""):
            domain_stats[dom]["likely"] += 1
        elif "Probable" in p.get("recommendation", ""):
            domain_stats[dom]["probable"] += 1
        elif "Similar" in p.get("recommendation", ""):
            domain_stats[dom]["similar"] += 1
        domain_stats[dom]["ds_ids"].add(p["dataset_a_id"])
        domain_stats[dom]["ds_ids"].add(p["dataset_b_id"])

    sorted_domains = sorted(domain_stats.items(), key=lambda x: -x[1]["total"])
    for row_idx, (dom, stats) in enumerate(sorted_domains, 2):
        row_data = [
            dom, stats["total"], stats["likely"],
            stats["probable"], stats["similar"], len(stats["ds_ids"]),
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_dom.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER

    dom_widths = [35, 12, 18, 12, 12, 18]
    for i, w in enumerate(dom_widths, 1):
        ws_dom.column_dimensions[get_column_letter(i)].width = w
    ws_dom.freeze_panes = "A2"

    # Save
    os.makedirs(OUTPUT_PATH.parent, exist_ok=True)
    wb.save(OUTPUT_PATH)

    print(f"\n{'='*60}")
    print("CONSOLIDATION WORKBOOK COMPLETE")
    print(f"{'='*60}")
    print(f"  Output: {OUTPUT_PATH}")
    print(f"  Tabs: Executive Summary, Consolidation Groups ({len(groups)} groups), "
          f"Pair Details ({min(len(pairs), 2000)} pairs), Domain Breakdown ({len(sorted_domains)} domains)")
    print(f"  Potential dataset reduction: {sum(len(g) - 1 for g in groups)} datasets")


if __name__ == "__main__":
    main()
