#!/usr/bin/env python3
"""
generate_owner_rollouts.py — Build per-owner cleanup spreadsheets for the
Domo governance rollout.

Each owner gets an Excel file with:
  - Their datasets flagged for review (stale, test/temp, orphaned)
  - Their dataflows flagged for review
  - Usage data: staleness, row count, lineage connections, domain classification
  - A "Decision" column (Keep / Remove) for them to fill in
  - A "Notes" column for justification

Also produces a summary manifest (rollout_manifest.csv) listing every owner,
their counts, email-ready status, and escalation info.
"""

import csv
import json
import os
import re
from collections import defaultdict
from datetime import datetime, timezone

import infer_descriptions
from generate_renames import apply_rename_rules
from generate_aggressive_renames import aggressive_restructure

# ── Dashboard impact data (loaded once at module level) ────────────────────
IMPACT_FILE = "output/dashboard_impact_report.csv"


def _load_dashboard_impact() -> dict[str, dict]:
    """Load dashboard impact data, keyed by dataset_id.

    Returns {dataset_id: {"card_count": N, "page_names": "P1; P2", "card_names": "C1; C2"}}
    """
    impact: dict[str, dict] = {}
    if not os.path.exists(IMPACT_FILE):
        return impact

    with open(IMPACT_FILE, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            ds_id = row.get("flagged_dataset_id", "")
            if ds_id not in impact:
                impact[ds_id] = {
                    "card_count": 0,
                    "card_names": [],
                    "page_names": set(),
                    "all_flagged_count": 0,
                }
            impact[ds_id]["card_count"] += 1
            card_title = row.get("card_title", "").strip()
            if card_title:
                impact[ds_id]["card_names"].append(card_title)
            pages = row.get("pages", "").strip()
            if pages:
                for p in pages.split("; "):
                    if p.strip():
                        impact[ds_id]["page_names"].add(p.strip())
            if row.get("all_datasets_flagged") == "Yes":
                impact[ds_id]["all_flagged_count"] += 1

    # Format for display
    result = {}
    for ds_id, data in impact.items():
        card_count = data["card_count"]
        page_list = sorted(data["page_names"])
        card_list = data["card_names"][:5]  # truncate to 5 card names
        suffix = f" (+{len(data['card_names']) - 5} more)" if len(data["card_names"]) > 5 else ""

        result[ds_id] = {
            "cards_affected": card_count,
            "pages_affected": "; ".join(page_list) if page_list else "",
            "card_details": "; ".join(card_list) + suffix if card_list else "",
        }
    return result


DASHBOARD_IMPACT = _load_dashboard_impact()

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Config ──────────────────────────────────────────────────────────────────
CACHE_FILE = ".cache/latest.json"
OUTPUT_DIR = "output/owner_rollouts"
MANIFEST_FILE = "output/rollout_manifest.csv"
DEADLINE = "May 1, 2026"
DOMO_INSTANCE = "gstv"  # → https://gstv.domo.com

# Decision dropdown options
DECISION_OPTIONS = ["Keep", "Remove", "Need Discussion", "Not My Dataset"]

# Staleness tiers (must match analytics.py)
STALENESS_TIERS = [
    (7, "Current"),
    (30, "Recent"),
    (90, "Aging"),
    (180, "Very Stale"),
    (365, "Dormant"),
]

# Domain classification rules (import from analytics)
import analytics

# ── Styles ──────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ABANDONED_FILL = PatternFill("solid", fgColor="FFC7CE")  # Red — abandoned
DORMANT_FILL = PatternFill("solid", fgColor="FFE0B2")     # Orange — dormant
STALE_FILL = PatternFill("solid", fgColor="FFF9C4")       # Yellow — very stale
TEST_FILL = PatternFill("solid", fgColor="E1BEE7")        # Purple — test/temp
ACTIVE_FILL = PatternFill("solid", fgColor="C8E6C9")      # Green — active (not flagged)
DECISION_FILL = PatternFill("solid", fgColor="E3F2FD")    # Light blue — decision column
RENAME_FILL = PatternFill("solid", fgColor="FFF3E0")      # Light orange — rename columns
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def get_staleness(days_old):
    if days_old is None:
        return "Unknown"
    for threshold, label in STALENESS_TIERS:
        if days_old <= threshold:
            return label
    return "Abandoned"


def parse_ts(ts):
    if not ts:
        return None
    try:
        dt = datetime.fromisoformat(str(ts).replace("Z", "+00:00"))
        # Ensure timezone-aware
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except (ValueError, TypeError):
        return None


def build_owner_data():
    """Load cache and build per-owner review lists."""
    with open(CACHE_FILE) as f:
        data = json.load(f)

    datasets = data["datasets"]
    dataflows = data["dataflows"]
    lineage = data["lineage"]

    now = datetime(2026, 4, 11, tzinfo=timezone.utc)

    # Build owner ID → name map
    id_to_name = {}
    for d in datasets:
        if d.get("owner_name") and d.get("owner_id"):
            id_to_name[d["owner_id"]] = d["owner_name"]

    # Build lineage sets
    ds_input_to_df = defaultdict(set)   # dataset_id → set of dataflow_names it feeds
    ds_output_of_df = defaultdict(set)  # dataset_id → set of dataflow_names that produce it
    df_in_lineage = set()

    for rec in lineage:
        ds_id = rec.get("dataset_id", "")
        df_id = rec.get("dataflow_id", "")
        df_name = rec.get("dataflow_name", "")
        direction = rec.get("direction", "")

        if df_id:
            df_in_lineage.add(str(df_id))

        if direction == "Input" and ds_id:
            ds_input_to_df[ds_id].add(df_name.strip())
        elif direction == "Output" and ds_id:
            ds_output_of_df[ds_id].add(df_name.strip())

    all_ds_in_lineage = set(ds_input_to_df.keys()) | set(ds_output_of_df.keys())

    # ── Process datasets ──
    owner_datasets = defaultdict(list)

    for ds in datasets:
        ds_id = ds.get("dataset_id", "")
        ds_name = ds.get("dataset_name", "")
        owner = ds.get("owner_name", "Unknown")
        data_current = parse_ts(ds.get("data_current_at", ""))
        created = parse_ts(ds.get("created_at", ""))
        updated = parse_ts(ds.get("updated_at", ""))
        days_stale = (now - data_current).days if data_current else None
        staleness = get_staleness(days_stale)
        has_lineage = ds_id in all_ds_in_lineage
        domain, dept = analytics._classify_domain(ds_name)
        row_count = ds.get("row_count", 0) or 0
        col_count = ds.get("column_count", 0) or 0

        # Determine if this should be flagged for review
        flag_reason = ""
        priority = 99

        if staleness == "Abandoned" and not has_lineage:
            flag_reason = "Abandoned — no data updates in 365+ days, no dataflow connections"
            priority = 1
        elif staleness == "Dormant" and not has_lineage:
            flag_reason = "Dormant — no data updates in 180+ days, no dataflow connections"
            priority = 2
        elif staleness == "Very Stale" and not has_lineage:
            flag_reason = "Very Stale — no data updates in 90+ days, no dataflow connections"
            priority = 3
        elif staleness == "Abandoned" and has_lineage:
            flag_reason = "Abandoned — 365+ days stale but connected to dataflow(s)"
            priority = 4
        elif staleness == "Dormant" and has_lineage:
            flag_reason = "Dormant — 180+ days stale but connected to dataflow(s)"
            priority = 5
        elif domain == "Test / Temp / Archive":
            flag_reason = "Named as test, temp, dev, or archive dataset"
            priority = 6
        elif staleness == "Unknown":
            flag_reason = "No data freshness timestamp available"
            priority = 7
        else:
            continue  # Healthy — don't include

        # Build downstream info
        feeds = sorted(ds_input_to_df.get(ds_id, set()))
        produced_by = sorted(ds_output_of_df.get(ds_id, set()))
        lineage_desc = ""
        if feeds:
            lineage_desc += f"Feeds: {'; '.join(feeds[:3])}"
            if len(feeds) > 3:
                lineage_desc += f" (+{len(feeds)-3} more)"
        if produced_by:
            if lineage_desc:
                lineage_desc += " | "
            lineage_desc += f"Produced by: {'; '.join(produced_by[:3])}"
            if len(produced_by) > 3:
                lineage_desc += f" (+{len(produced_by)-3} more)"

        domo_url = f"https://{DOMO_INSTANCE}.domo.com/datasources/{ds_id}/details/overview"

        # Description: existing or inferred
        existing_desc = ds.get("description", "").strip()
        if not existing_desc:
            lineage_info = {
                'feeds': sorted(ds_input_to_df.get(ds_id, set())),
                'produced_by': sorted(ds_output_of_df.get(ds_id, set())),
            }
            suggested_desc = infer_descriptions.infer_dataset_description(
                ds, domain, dept, lineage_info
            ) or ""
        else:
            suggested_desc = ""

        description = existing_desc or suggested_desc

        # Compute rename suggestions
        proposed_name, cons_changes = apply_rename_rules(ds_name)
        if not cons_changes or proposed_name == ds_name.strip():
            proposed_name = ""  # No conservative rename needed

        restructured_name, restr_changed = aggressive_restructure(
            ds_name, domain, dept, "dataset"
        )
        if not restr_changed or restructured_name == ds_name.strip():
            restructured_name = ""
        # Don't show restructured if same as conservative
        if restructured_name and restructured_name == proposed_name:
            restructured_name = ""

        # Dashboard impact
        ds_impact = DASHBOARD_IMPACT.get(ds_id, {})
        cards_affected = ds_impact.get("cards_affected", 0)
        pages_affected = ds_impact.get("pages_affected", "")
        card_details = ds_impact.get("card_details", "")

        record = {
            "Dataset Name": ds_name,
            "Type": "Dataset",
            "Domo Link": domo_url,
            "Description": description,
            "Proposed Name": proposed_name,
            "Restructured Name": restructured_name,
            "Cards Affected": cards_affected if cards_affected else "",
            "Pages Affected": pages_affected,
            "Domain": domain,
            "Department": dept,
            "Staleness": staleness,
            "Days Since Data Update": days_stale if days_stale is not None else "N/A",
            "Last Data Update": data_current.strftime("%Y-%m-%d") if data_current else "Never",
            "Created": created.strftime("%Y-%m-%d") if created else "",
            "Row Count": row_count,
            "Column Count": col_count,
            "Has Lineage": "Yes" if has_lineage else "No",
            "Lineage Details": lineage_desc,
            "Flag Reason": flag_reason,
            "Decision": "",
            "Notes / Justification": "",
            "_priority": priority,
        }

        owner_datasets[owner].append(record)

    # ── Process dataflows ──
    owner_dataflows = defaultdict(list)

    for df in dataflows:
        df_id = str(df.get("dataflow_id", ""))
        df_name = df.get("dataflow_name", "").strip()
        owner_id = df.get("owner_id")
        # Resolve name: dataflow API doesn't return owner_name, so use ID lookup
        # Note: dataflow owner_ids are strings, dataset owner_ids are ints
        raw_name = (df.get("owner_name") or "").strip()
        if raw_name:
            owner = raw_name
        else:
            int_id = int(owner_id) if str(owner_id).isdigit() else owner_id
            owner = id_to_name.get(owner_id, id_to_name.get(int_id, None))
            if not owner:
                # Unknown owner — likely former employee; assign to Data team
                owner = "Data Team (Former Owner)"
        last_exec = parse_ts(df.get("last_execution_date", ""))
        last_updated = parse_ts(df.get("last_updated_date", ""))
        days_since_exec = (now - last_exec).days if last_exec else None
        staleness = get_staleness(days_since_exec)
        has_lineage = df_id in df_in_lineage
        domain, dept = analytics._classify_domain(df_name)
        df_type = df.get("dataflow_type", "")
        status = df.get("status", "")
        input_count = df.get("input_count", 0) or 0
        output_count = df.get("output_count", 0) or 0

        flag_reason = ""
        priority = 99

        if staleness == "Abandoned":
            flag_reason = "Abandoned — not executed in 365+ days"
            priority = 1
        elif staleness == "Dormant":
            flag_reason = "Dormant — not executed in 180+ days"
            priority = 3
        elif staleness == "Very Stale":
            flag_reason = "Very Stale — not executed in 90+ days"
            priority = 5
        elif domain == "Test / Temp / Archive":
            flag_reason = "Named as test, temp, dev, or archive dataflow"
            priority = 6
        else:
            continue

        domo_url = f"https://{DOMO_INSTANCE}.domo.com/datacenter/dataflows/{df_id}/details#datasets"

        # Description: existing or inferred
        existing_desc = df.get("description", "").strip()
        if not existing_desc:
            suggested_desc = infer_descriptions.infer_dataflow_description(
                df, domain
            ) or ""
        else:
            suggested_desc = ""

        description = existing_desc or suggested_desc

        # Compute rename suggestions
        proposed_name, cons_changes = apply_rename_rules(df_name)
        if not cons_changes or proposed_name == df_name.strip():
            proposed_name = ""

        restructured_name, restr_changed = aggressive_restructure(
            df_name, domain, dept, "dataflow"
        )
        if not restr_changed or restructured_name == df_name.strip():
            restructured_name = ""
        if restructured_name and restructured_name == proposed_name:
            restructured_name = ""

        record = {
            "Dataflow Name": df_name,
            "Type": "Dataflow",
            "Domo Link": domo_url,
            "Description": description,
            "Proposed Name": proposed_name,
            "Restructured Name": restructured_name,
            "Dataflow Type": df_type,
            "Domain": domain,
            "Department": dept,
            "Staleness": staleness,
            "Days Since Last Execution": days_since_exec if days_since_exec is not None else "N/A",
            "Last Executed": last_exec.strftime("%Y-%m-%d") if last_exec else "Never",
            "Status": status,
            "Input Datasets": input_count,
            "Output Datasets": output_count,
            "Has Lineage": "Yes" if has_lineage else "No",
            "Flag Reason": flag_reason,
            "Decision": "",
            "Notes / Justification": "",
            "_priority": priority,
        }

        owner_dataflows[owner].append(record)

    return owner_datasets, owner_dataflows, id_to_name


def write_owner_workbook(owner_name: str, ds_records: list, df_records: list):
    """Write one Excel workbook for a single owner."""
    wb = Workbook()

    # ── Instructions tab ──
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    ws_instr.sheet_properties.tabColor = "1F4E79"

    instructions = [
        ("Domo Cleanup Review", None),
        ("", None),
        (f"Owner: {owner_name}", None),
        (f"Date: April 11, 2026", None),
        (f"Response Deadline: {DEADLINE}", None),
        ("", None),
        ("WHY THIS MATTERS", None),
        ("We are preparing our Domo environment for AI-powered analytics and automated", None),
        ("reporting. To enable these capabilities, we need a clean, well-organized data", None),
        ("environment. Unused and stale datasets create noise that degrades AI accuracy", None),
        ("and increases our Domo costs.", None),
        ("", None),
        ("WHAT WE NEED FROM YOU", None),
        ("1. Review the Datasets and Dataflows tabs in this workbook", None),
        ("2. Click the Domo Link to open each item and verify its purpose", None),
        ("3. Use the Decision dropdown: Keep, Remove, Need Discussion, or Not My Dataset", None),
        ("4. If keeping, provide a brief justification in the Notes column", None),
        ("5. Review the Proposed Name and Restructured Name columns for rename suggestions", None),
        (f"6. Return this spreadsheet by {DEADLINE}", None),
        ("", None),
        ("RENAME COLUMNS", None),
        ("Two rename suggestion columns are provided for items you keep:", None),
        ("  - Proposed Name: Conservative fix (extensions, brackets, casing only)", None),
        ("  - Restructured Name: Full convention compliance (Domain - Description - Qualifier)", None),
        ("These are suggestions — review and override as needed.", None),
        ("", None),
        ("IMPORTANT", None),
        (f"Items without a response by {DEADLINE} will be removed.", None),
        ("If you choose to keep a dataset, you will be responsible for:", None),
        ("  - Maintaining the data pipeline and ensuring freshness", None),
        ("  - Providing column definitions in a follow-up request", None),
        ("  - Identifying a backup contact who understands the data", None),
        ("", None),
        ("QUESTIONS?", None),
        ("Contact Garrett Kohler or Sriram Vepuri on the Data team.", None),
    ]

    for row_idx, (text, _) in enumerate(instructions, 1):
        cell = ws_instr.cell(row=row_idx, column=1, value=text)
        if row_idx == 1:
            cell.font = Font(bold=True, size=16, color="1F4E79")
        elif text in ("WHY THIS MATTERS", "WHAT WE NEED FROM YOU", "RENAME COLUMNS", "IMPORTANT", "QUESTIONS?"):
            cell.font = Font(bold=True, size=12, color="1F4E79")
        elif text.startswith(f"Items without a response by"):
            cell.font = Font(bold=True, color="C00000")
        else:
            cell.font = Font(size=11)

    ws_instr.column_dimensions["A"].width = 85

    # ── Datasets tab ──
    if ds_records:
        ds_records.sort(key=lambda r: r["_priority"])
        ws_ds = wb.create_sheet("Datasets")
        ws_ds.sheet_properties.tabColor = "C00000"

        ds_cols = [
            "Dataset Name", "Domo Link", "Description",
            "Proposed Name", "Restructured Name",
            "Cards Affected", "Pages Affected",
            "Domain", "Department",
            "Staleness", "Days Since Data Update", "Last Data Update", "Created",
            "Row Count", "Column Count", "Has Lineage", "Lineage Details",
            "Flag Reason", "Decision", "Notes / Justification",
        ]
        _write_review_sheet(ws_ds, ds_records, ds_cols)

    # ── Dataflows tab ──
    if df_records:
        df_records.sort(key=lambda r: r["_priority"])
        ws_df = wb.create_sheet("Dataflows")
        ws_df.sheet_properties.tabColor = "E65100"

        df_cols = [
            "Dataflow Name", "Domo Link", "Description",
            "Proposed Name", "Restructured Name",
            "Dataflow Type", "Domain",
            "Department", "Staleness", "Days Since Last Execution", "Last Executed",
            "Status", "Input Datasets", "Output Datasets", "Has Lineage",
            "Flag Reason", "Decision", "Notes / Justification",
        ]
        _write_review_sheet(ws_df, df_records, df_cols)

    # ── Summary tab ──
    ws_sum = wb.create_sheet("Summary", 0)  # Insert at front
    ws_sum.sheet_properties.tabColor = "2E7D32"
    wb.active = ws_sum

    ds_by_staleness = defaultdict(int)
    for r in ds_records:
        ds_by_staleness[r["Staleness"]] += 1

    df_by_staleness = defaultdict(int)
    for r in df_records:
        df_by_staleness[r["Staleness"]] += 1

    summary_data = [
        ("YOUR CLEANUP REVIEW SUMMARY", ""),
        ("", ""),
        ("Datasets Flagged for Review", len(ds_records)),
        ("  Abandoned (365+ days)", ds_by_staleness.get("Abandoned", 0)),
        ("  Dormant (180-365 days)", ds_by_staleness.get("Dormant", 0)),
        ("  Very Stale (90-180 days)", ds_by_staleness.get("Very Stale", 0)),
        ("  Test / Temp / Archive", sum(1 for r in ds_records if "Test / Temp" in r.get("Domain", ""))),
        ("  Unknown freshness", ds_by_staleness.get("Unknown", 0)),
        ("", ""),
        ("Dataflows Flagged for Review", len(df_records)),
        ("  Abandoned (365+ days)", df_by_staleness.get("Abandoned", 0)),
        ("  Dormant (180-365 days)", df_by_staleness.get("Dormant", 0)),
        ("  Very Stale (90-180 days)", df_by_staleness.get("Very Stale", 0)),
        ("  Test / Temp / Archive", sum(1 for r in df_records if "Test / Temp" in r.get("Domain", ""))),
        ("", ""),
        ("Total Items for Review", len(ds_records) + len(df_records)),
        ("", ""),
        ("Datasets Powering Cards/Dashboards", sum(1 for r in ds_records if r.get("Cards Affected"))),
        ("", ""),
        (f"DEADLINE: {DEADLINE}", ""),
        ("Items without a response will be removed.", ""),
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_a = ws_sum.cell(row=row_idx, column=1, value=label)
        cell_b = ws_sum.cell(row=row_idx, column=2, value=value if value != "" else None)

        if row_idx == 1:
            cell_a.font = Font(bold=True, size=14, color="1F4E79")
        elif label.startswith("  "):
            cell_a.font = Font(size=11, color="666666")
            cell_a.alignment = Alignment(indent=2)
        elif label.startswith("DEADLINE"):
            cell_a.font = Font(bold=True, size=12, color="C00000")
        elif label == "Items without a response will be removed.":
            cell_a.font = Font(bold=True, color="C00000")
        elif value and isinstance(value, int) and value > 0:
            cell_a.font = Font(bold=True, size=11)
            cell_b.font = Font(bold=True, size=11)

    ws_sum.column_dimensions["A"].width = 45
    ws_sum.column_dimensions["B"].width = 15

    # Save
    safe_name = re.sub(r'[^\w\-]', '_', owner_name)
    filepath = os.path.join(OUTPUT_DIR, f"cleanup_review_{safe_name}.xlsx")
    wb.save(filepath)
    return filepath


def _write_review_sheet(ws, records, columns):
    """Write a formatted review sheet with header, data, and conditional formatting."""
    # Header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Find column indices for special handling
    link_col_idx = None
    decision_col_idx = None
    rename_col_indices = set()
    cards_affected_col_idx = None
    for col_idx, col_name in enumerate(columns, 1):
        if col_name == "Domo Link":
            link_col_idx = col_idx
        elif col_name == "Decision":
            decision_col_idx = col_idx
        elif col_name in ("Proposed Name", "Restructured Name"):
            rename_col_indices.add(col_idx)
        elif col_name == "Cards Affected":
            cards_affected_col_idx = col_idx

    # Data rows
    for row_idx, record in enumerate(records, 2):
        staleness = record.get("Staleness", "")

        for col_idx, col_name in enumerate(columns, 1):
            value = record.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Domo Link column: make it a clickable hyperlink
            if col_idx == link_col_idx and value:
                cell.value = "Open in Domo"
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")
                cell.fill = PatternFill()  # No staleness coloring on link
                continue

            cell.value = value

            # Color code by staleness
            if col_name in ("Decision", "Notes / Justification"):
                cell.fill = DECISION_FILL
                cell.font = Font(size=11)
            elif col_idx in rename_col_indices:
                if value:  # Only color if there's a rename suggestion
                    cell.fill = RENAME_FILL
                    cell.font = Font(size=10, italic=True)
            elif col_idx == cards_affected_col_idx and value and value > 0:
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
                cell.font = Font(bold=True, color="C00000")
            elif staleness == "Abandoned":
                cell.fill = ABANDONED_FILL
            elif staleness == "Dormant":
                cell.fill = DORMANT_FILL
            elif staleness == "Very Stale":
                cell.fill = STALE_FILL
            elif "Test / Temp" in record.get("Domain", record.get("Flag Reason", "")):
                cell.fill = TEST_FILL

    # Add dropdown validation to Decision column
    if decision_col_idx and len(records) > 0:
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(DECISION_OPTIONS)}"',
            allow_blank=True,
        )
        dv.error = "Please select from the dropdown: Keep, Remove, Need Discussion, or Not My Dataset"
        dv.errorTitle = "Invalid Decision"
        dv.prompt = "Select your decision for this item"
        dv.promptTitle = "Decision"
        decision_letter = get_column_letter(decision_col_idx)
        dv.add(f"{decision_letter}2:{decision_letter}{len(records) + 1}")
        ws.add_data_validation(dv)

    # Column widths
    col_widths = {
        "Dataset Name": 45, "Dataflow Name": 45,
        "Domo Link": 15, "Description": 50,
        "Domain": 22, "Department": 18, "Staleness": 14,
        "Days Since Data Update": 12, "Days Since Last Execution": 12,
        "Last Data Update": 13, "Last Executed": 13, "Created": 13,
        "Row Count": 12, "Column Count": 10, "Input Datasets": 10, "Output Datasets": 10,
        "Has Lineage": 10, "Lineage Details": 40, "Flag Reason": 45,
        "Proposed Name": 45, "Restructured Name": 45,
        "Cards Affected": 13, "Pages Affected": 30,
        "Decision": 18, "Notes / Justification": 35,
        "Dataflow Type": 14, "Status": 12, "Type": 10,
    }
    for col_idx, col_name in enumerate(columns, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_widths.get(col_name, 15)

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(records) + 1}"


def write_manifest(owner_datasets, owner_dataflows, id_to_name):
    """Write rollout_manifest.csv with per-owner summary."""
    all_owners = sorted(set(list(owner_datasets.keys()) + list(owner_dataflows.keys())))

    rows = []
    for owner in all_owners:
        ds = owner_datasets.get(owner, [])
        dfs = owner_dataflows.get(owner, [])

        ds_abandoned = sum(1 for r in ds if r["Staleness"] == "Abandoned")
        ds_dormant = sum(1 for r in ds if r["Staleness"] == "Dormant")
        ds_stale = sum(1 for r in ds if r["Staleness"] == "Very Stale")
        ds_test = sum(1 for r in ds if "Test / Temp" in r.get("Domain", ""))

        df_abandoned = sum(1 for r in dfs if r["Staleness"] == "Abandoned")
        df_dormant = sum(1 for r in dfs if r["Staleness"] == "Dormant")

        rows.append({
            "Owner": owner,
            "Total Items for Review": len(ds) + len(dfs),
            "Datasets Flagged": len(ds),
            "DS Abandoned": ds_abandoned,
            "DS Dormant": ds_dormant,
            "DS Very Stale": ds_stale,
            "DS Test/Temp": ds_test,
            "Dataflows Flagged": len(dfs),
            "DF Abandoned": df_abandoned,
            "DF Dormant": df_dormant,
            "Spreadsheet": f"cleanup_review_{re.sub(r'[^\\w-]', '_', owner)}.xlsx",
            "Deadline": DEADLINE,
            "Escalation": "Garrett Kohler / Sriram Vepuri",
        })

    rows.sort(key=lambda r: -r["Total Items for Review"])

    with open(MANIFEST_FILE, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)

    return rows


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("Building per-owner review data...")
    owner_datasets, owner_dataflows, id_to_name = build_owner_data()

    all_owners = sorted(set(list(owner_datasets.keys()) + list(owner_dataflows.keys())))
    print(f"  Found {len(all_owners)} owners with items to review\n")

    total_ds = 0
    total_df = 0

    for owner in all_owners:
        ds = owner_datasets.get(owner, [])
        dfs = owner_dataflows.get(owner, [])
        total_ds += len(ds)
        total_df += len(dfs)

        filepath = write_owner_workbook(owner, ds, dfs)
        print(f"  {owner:25s}: {len(ds):>4d} datasets, {len(dfs):>4d} dataflows → {os.path.basename(filepath)}")

    print(f"\n  TOTAL: {total_ds} datasets + {total_df} dataflows across {len(all_owners)} owners")

    # Write manifest
    manifest = write_manifest(owner_datasets, owner_dataflows, id_to_name)
    print(f"\nManifest: {MANIFEST_FILE}")
    print(f"Spreadsheets: {OUTPUT_DIR}/")

    # Print summary table
    print(f"\n{'='*80}")
    print(f"{'Owner':25s} | {'Datasets':>8s} | {'Dataflows':>9s} | {'Total':>5s}")
    print(f"{'-'*25}-+-{'-'*8}-+-{'-'*9}-+-{'-'*5}")
    for r in manifest:
        print(f"{r['Owner']:25s} | {r['Datasets Flagged']:>8d} | {r['Dataflows Flagged']:>9d} | {r['Total Items for Review']:>5d}")


if __name__ == "__main__":
    main()
