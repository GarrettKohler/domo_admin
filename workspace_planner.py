"""Generate the Workspace Planning & Naming Conventions workbook.

Produces a standalone Excel workbook for team review covering:
- Workspace summary with health metrics
- Dataset-to-workspace assignments
- Naming convention standards
- Proposed renames for datasets that violate conventions
"""

import json
import re
import logging
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from analytics import build_domain_map, build_dataset_lineage_analysis, build_cleanup_candidates
from extractors import load_cache

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="3E5170", end_color="3E5170", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)
SECTION_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)

HEALTH_FILLS = {
    "Healthy":   PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid"),
    "Fair":      PatternFill(start_color="F9A825", end_color="F9A825", fill_type="solid"),
    "Poor":      PatternFill(start_color="E65100", end_color="E65100", fill_type="solid"),
    "Critical":  PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid"),
}
HEALTH_FONT_WHITE = Font(name="Arial", size=10, color="FFFFFF")
HEALTH_FONT_DARK = Font(name="Arial", size=10, color="000000")

WORKSPACE_FILLS = {
    "Sites & Locations":       PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid"),
    "Proof of Play":           PatternFill(start_color="00695C", end_color="00695C", fill_type="solid"),
    "Impressions":             PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid"),
    "Transactions":            PatternFill(start_color="4527A0", end_color="4527A0", fill_type="solid"),
    "Revenue & Monetization":  PatternFill(start_color="AD1457", end_color="AD1457", fill_type="solid"),
    "Programmatic Operations": PatternFill(start_color="E65100", end_color="E65100", fill_type="solid"),
    "Programmatic":            PatternFill(start_color="BF360C", end_color="BF360C", fill_type="solid"),
    "RPA":                     PatternFill(start_color="00838F", end_color="00838F", fill_type="solid"),
    "Traffic Instructions":    PatternFill(start_color="558B2F", end_color="558B2F", fill_type="solid"),
    "Managed Services":        PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid"),
    "Campaigns & Delivery":    PatternFill(start_color="C62828", end_color="C62828", fill_type="solid"),
    "Salesforce / CRM":        PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid"),
    "Monitoring & Governance": PatternFill(start_color="37474F", end_color="37474F", fill_type="solid"),
    "Site Analytics":          PatternFill(start_color="00695C", end_color="00695C", fill_type="solid"),
    "Engineering":             PatternFill(start_color="455A64", end_color="455A64", fill_type="solid"),
    "Test / Temp / Archive":   PatternFill(start_color="757575", end_color="757575", fill_type="solid"),
    "Other / Unclassified":    PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid"),
}
WORKSPACE_FONT = Font(name="Arial", size=10, color="FFFFFF")

STALENESS_FILLS = {
    "Active":    PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid"),
    "Stale":     PatternFill(start_color="F9A825", end_color="F9A825", fill_type="solid"),
    "Very Stale": PatternFill(start_color="E65100", end_color="E65100", fill_type="solid"),
    "Dormant":   PatternFill(start_color="BF360C", end_color="BF360C", fill_type="solid"),
    "Abandoned": PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid"),
    "Unknown":   PatternFill(start_color="757575", end_color="757575", fill_type="solid"),
}
STALENESS_FONT = Font(name="Arial", size=10, color="FFFFFF")


# ---------------------------------------------------------------------------
# Naming convention rules & logic
# ---------------------------------------------------------------------------

# Subsection assignment rules — maps (domain, name_pattern) to subsection
SUBSECTION_RULES = [
    # Sites & Locations subsections
    ("Sites & Locations", r"(?i)tdlinx", "TDLinx"),
    ("Sites & Locations", r"(?i)dxp|dxpromote", "DXP / DXPromote Conversions"),
    ("Sites & Locations", r"(?i)site.*status|status.*history", "Site Status History"),
    ("Sites & Locations", r"(?i)master.?station", "Master Station List"),
    ("Sites & Locations", r"(?i)mongo|gbase", "Source Systems (Mongo / GBase)"),
    ("Sites & Locations", r"(?i)market.?plan", "Market Plan"),
    ("Sites & Locations", r"(?i)noc\b|full.?status|jupiter|enterprise.*iotv|iotv|dispenser", "NOC Operations"),
    ("Sites & Locations", r"(?i)dma.*gtvid", "Operational"),
    ("Sites & Locations", r"(?i)sites.*base|site.*list|site.*config", "Base Site Data"),

    # Revenue & Monetization subsections
    ("Revenue & Monetization", r"(?i)clawback", "Clawback"),
    ("Revenue & Monetization", r"(?i)revenue.?share|rev.?share", "Revenue Share"),
    ("Revenue & Monetization", r"(?i)rent.?payment", "Rent Payments"),
    ("Revenue & Monetization", r"(?i)retailer.?campaign", "Retailer Campaign Revenue"),
    ("Revenue & Monetization", r"(?i)cpm.?floor", "CPM & Pricing"),
    ("Revenue & Monetization", r"(?i)pluto", "Pluto"),
    ("Revenue & Monetization", r"(?i)vistar|place.?exchange|px|broadsign|hivestack|magnite|ssp|programmatic.*rev|exchange.*rev", "Programmatic Revenue"),
    ("Revenue & Monetization", r"(?i)invoice", "Invoicing"),

    # Programmatic Operations subsections
    ("Programmatic Operations", r"(?i)vistar.*(diagnos|fill)", "Vistar Diagnostics"),
    ("Programmatic Operations", r"(?i)vistar.*avail|venue.*avail", "Vistar Venue Availability"),
    ("Programmatic Operations", r"(?i)auction|bid", "Auction & Bid Data"),
    ("Programmatic Operations", r"(?i)sell.?through", "Sell-Through Analysis"),
    ("Programmatic Operations", r"(?i)place.?exchange", "Partner - PlaceExchange"),
    ("Programmatic Operations", r"(?i)magnite", "Partner - Magnite"),
    ("Programmatic Operations", r"(?i)hivestack", "Partner - Hivestack"),
    ("Programmatic Operations", r"(?i)broadsign", "Partner - Broadsign"),

    # Programmatic subsections
    ("Programmatic", r"(?i)vistar", "Partner - Vistar"),
    ("Programmatic", r"(?i)place.?exchange|px", "Partner - PlaceExchange"),
    ("Programmatic", r"(?i)broadsign", "Partner - Broadsign"),
    ("Programmatic", r"(?i)hivestack", "Partner - Hivestack"),
    ("Programmatic", r"(?i)magnite", "Partner - Magnite"),

    # Proof of Play subsections
    ("Proof of Play", r"(?i)gilbarco.*ics|ics.*gilbarco", "Gilbarco ICS"),
    ("Proof of Play", r"(?i)dover|dxpromote", "Dover DXPromote"),
    ("Proof of Play", r"(?i)speedway", "Speedway"),
    ("Proof of Play", r"(?i)simpli", "Partner - Simpli.fi"),

    # Transactions subsections
    ("Transactions", r"(?i)gilbarco", "Gilbarco"),
    ("Transactions", r"(?i)speedway", "Speedway"),
    ("Transactions", r"(?i)wayne", "Wayne"),
    ("Transactions", r"(?i)dover", "Dover"),
    ("Transactions", r"(?i)patch", "Patches"),
    ("Transactions", r"(?i)monitor", "Monitoring"),
    ("Transactions", r"(?i)unvalidated", "Unvalidated"),

    # Managed Services subsections
    ("Managed Services", r"(?i)casey", "Casey's"),
    ("Managed Services", r"(?i)speedway", "Speedway"),
    ("Managed Services", r"(?i)circle.?k", "Circle K"),
    ("Managed Services", r"(?i)kwik", "Kwik Trip"),
    ("Managed Services", r"(?i)pilot", "Pilot"),
    ("Managed Services", r"(?i)sheetz", "Sheetz"),
    ("Managed Services", r"(?i)holiday", "Holiday Station"),
    ("Managed Services", r"(?i)wawa", "Wawa"),
    ("Managed Services", r"(?i)marathon", "Marathon"),

    # Monitoring & Governance subsections
    ("Monitoring & Governance", r"(?i)domostats|governance", "DomoStats / Governance"),
    ("Monitoring & Governance", r"(?i)data.?quality|data.?check", "Data Quality Checks"),
    ("Monitoring & Governance", r"(?i)pmar", "PMAR"),
]


def _assign_subsection(domain: str, dataset_name: str) -> str:
    """Assign a subsection within a workspace based on dataset name."""
    for rule_domain, pattern, subsection in SUBSECTION_RULES:
        if rule_domain == domain and re.search(pattern, dataset_name):
            return subsection
    return "General"


# ---------------------------------------------------------------------------
# Naming convention analysis
# ---------------------------------------------------------------------------

NAMING_ISSUES = [
    # (issue_id, description, pattern_to_detect, proposed_fix_description)
    ("bracket_prefix", "Uses bracketed prefix instead of standard",
     r"^\[(?:Production|PROD|production|prod)\]", "Replace with 'PROD -'"),
    ("bracket_dev", "Uses bracketed DEV prefix",
     r"^\[(?:DEV|Dev|dev)\]", "Replace with 'DEV -'"),
    ("bracket_test", "Uses bracketed TEST prefix",
     r"^\[(?:TEST|Test|test)\]", "Replace with 'TEST -'"),
    ("bracket_dnu", "Uses [DNU] instead of DEPRECATED",
     r"\[DNU\]", "Replace with 'DEPRECATED -'"),
    ("flag_deletion", "Contains 'Flag for Deletion' tag",
     r"(?i)flag.?for.?deletion", "Add to cleanup queue, prefix with 'DEPRECATED -'"),
    ("file_extension", "Contains file extension in name",
     r"\.(xlsx|csv|json|txt)\s*$", "Remove file extension from name"),
    ("underscore_sep", "Uses underscores as separators",
     r"^[a-zA-Z]+_[a-zA-Z]+_", "Replace underscores with ' - '"),
    ("vw_prefix", "Uses 'vw_' database-style prefix",
     r"^vw_", "Replace with 'View - '"),
    ("view_of_view", "Nested 'View of View of'",
     r"(?i)^view of view of", "Simplify to single 'View - '"),
    ("view_of", "Uses 'View of' instead of 'View -'",
     r"(?i)^view of\s", "Replace with 'View - '"),
    ("editable_view", "Uses 'Editable DataSet for View of'",
     r"(?i)^editable\s+dataset\s+for\s+view\s+of", "Replace with 'Editable View - '"),
    ("double_space", "Contains double spaces",
     r"  ", "Remove extra spaces"),
    ("month_name_date", "Uses month names instead of ISO dates",
     r"(?i)\b(?:january|february|march|april|may|june|july|august|september|october|november|december)\s+20\d{2}\b", "Use YYYY-MM format"),
    ("mmddyy_date", "Uses MMDDYY date format",
     r"\b\d{6}\b(?![-\w])", "Use YYYY-MM-DD format"),
    ("no_dash_sep", "Uses inconsistent separators (periods, parens)",
     r"(?<!\d)\.(?!\d)|(?<!\()\((?!.*\).*-)", None),  # Informational only
    ("trailing_space", "Has trailing whitespace",
     r"\s+$", "Trim trailing spaces"),
    ("abbrev_network", "Uses abbreviated network name (GVR instead of Gilbarco)",
     r"\bGVR\b", "Use full name 'Gilbarco'"),
]


def _detect_naming_issues(name: str) -> list[dict[str, str]]:
    """Detect naming convention violations in a dataset name."""
    issues = []
    for issue_id, description, pattern, fix in NAMING_ISSUES:
        if fix is None:
            continue  # Informational only
        if re.search(pattern, name):
            issues.append({
                "issue": issue_id,
                "description": description,
                "fix": fix,
            })
    return issues


def _propose_rename(name: str) -> str:
    """Apply all naming convention rules to produce a proposed name."""
    proposed = name

    # 1. Bracketed prefixes
    proposed = re.sub(r"^\[(?:Production|PROD|production|prod)\]\s*-?\s*", "PROD - ", proposed)
    proposed = re.sub(r"^\[(?:DEV|Dev|dev)\]\s*-?\s*", "DEV - ", proposed)
    proposed = re.sub(r"^\[(?:TEST|Test|test)\]\s*-?\s*", "TEST - ", proposed)
    proposed = re.sub(r"\[DNU\]\s*-?\s*", "DEPRECATED - ", proposed)
    proposed = re.sub(r"(?i)\[flag.?for.?deletion\]\s*", "DEPRECATED - ", proposed)

    # 2. File extensions
    proposed = re.sub(r"\.(xlsx|csv|json|txt)\s*$", "", proposed, flags=re.IGNORECASE)

    # 3. vw_ prefix
    proposed = re.sub(r"^vw_", "View - ", proposed)

    # 4. View of View of → View -
    proposed = re.sub(r"(?i)^view\s+of\s+view\s+of\s+", "View - ", proposed)

    # 5. View of → View -
    proposed = re.sub(r"(?i)^view\s+of\s+", "View - ", proposed)

    # 6. Editable DataSet for View of
    proposed = re.sub(r"(?i)^editable\s+dataset\s+for\s+view\s+of\s+", "Editable View - ", proposed)

    # 7. Abbreviated network names
    proposed = re.sub(r"\bGVR\b", "Gilbarco", proposed)

    # 8. Double spaces
    proposed = re.sub(r"  +", " ", proposed)

    # 9. Trailing spaces
    proposed = proposed.strip()

    return proposed


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _apply_header(ws) -> None:
    """Apply standard header formatting and freeze top row."""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"


def _auto_fit(ws, max_width: int = 60) -> None:
    """Auto-fit column widths."""
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), 1):
        max_len = 0
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, min(len(val), max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)


def _write_workspace_summary(ws, domain_map: list[dict[str, Any]]) -> None:
    """Tab 1: Workspace summary with health metrics and subsections."""
    # Aggregate stats
    stats = defaultdict(lambda: {
        "count": 0, "active": 0, "stale": 0, "very_stale": 0,
        "dormant": 0, "abandoned": 0, "unknown": 0,
        "sources": 0, "sinks": 0, "passthrough": 0, "orphans": 0,
        "total_rows": 0, "subsections": defaultdict(int),
    })

    for rec in domain_map:
        d = stats[rec["domain"]]
        d["count"] += 1
        staleness_key = rec["staleness"].lower().replace(" ", "_")
        if staleness_key in d:
            d[staleness_key] += 1
        role_key = {"Source": "sources", "Sink": "sinks", "Pass-through": "passthrough", "Orphan": "orphans"}.get(rec["role"], "orphans")
        d[role_key] += 1
        d["total_rows"] += rec.get("row_count", 0) or 0
        subsection = _assign_subsection(rec["domain"], rec["dataset_name"])
        d["subsections"][subsection] += 1

    columns = [
        "Workspace", "Total Datasets", "Active", "Stale", "Very Stale",
        "Dormant", "Abandoned", "Health %", "Health Rating",
        "Sources", "Sinks", "Pass-through", "Orphans", "Subsections",
    ]

    for col_idx, col in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col)

    # Sort by count descending, but put cleanup/unclassified at bottom
    sort_order = {"Test / Temp / Archive": 998, "Other / Unclassified": 999}
    sorted_domains = sorted(stats.keys(), key=lambda d: (sort_order.get(d, 0), -stats[d]["count"]))

    health_col = columns.index("Health Rating") + 1
    workspace_col = 1

    for row_idx, domain in enumerate(sorted_domains, 2):
        s = stats[domain]
        health_pct = s["active"] / s["count"] * 100 if s["count"] > 0 else 0
        if health_pct >= 70:
            health_rating = "Healthy"
        elif health_pct >= 50:
            health_rating = "Fair"
        elif health_pct >= 30:
            health_rating = "Poor"
        else:
            health_rating = "Critical"

        subsection_str = " | ".join(
            f"{name} ({count})" for name, count
            in sorted(s["subsections"].items(), key=lambda x: -x[1])
        )

        row_data = [
            domain, s["count"], s["active"], s["stale"], s["very_stale"],
            s["dormant"], s["abandoned"], f"{health_pct:.0f}%", health_rating,
            s["sources"], s["sinks"], s["passthrough"], s["orphans"],
            subsection_str,
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            if col_idx == len(row_data):  # Subsections column
                cell.alignment = WRAP_ALIGNMENT

        # Color workspace name
        ws_cell = ws.cell(row=row_idx, column=workspace_col)
        if domain in WORKSPACE_FILLS:
            ws_cell.fill = WORKSPACE_FILLS[domain]
            ws_cell.font = WORKSPACE_FONT

        # Color health rating
        h_cell = ws.cell(row=row_idx, column=health_col)
        if health_rating in HEALTH_FILLS:
            h_cell.fill = HEALTH_FILLS[health_rating]
            if health_rating == "Fair":
                h_cell.font = HEALTH_FONT_DARK
            else:
                h_cell.font = HEALTH_FONT_WHITE

    _apply_header(ws)
    _auto_fit(ws)
    # Make subsections column wider
    ws.column_dimensions[get_column_letter(len(columns))].width = 80


def _write_dataset_assignments(ws, domain_map: list[dict[str, Any]]) -> None:
    """Tab 2: Every dataset with workspace, subsection, role, staleness."""
    columns = [
        "dataset_name", "dataset_id", "workspace", "subsection",
        "role", "staleness", "days_since_update", "row_count",
        "owner_name", "feeds_dataflow_count", "fed_by_dataflow_count",
    ]

    for col_idx, col in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col)

    workspace_col = columns.index("workspace") + 1
    staleness_col = columns.index("staleness") + 1

    # Enrich with subsection
    records = []
    for rec in domain_map:
        r = dict(rec)
        r["workspace"] = rec["domain"]
        r["subsection"] = _assign_subsection(rec["domain"], rec["dataset_name"])
        records.append(r)

    # Sort by workspace, subsection, dataset name
    records.sort(key=lambda r: (r["workspace"].lower(), r["subsection"].lower(), r["dataset_name"].lower()))

    for row_idx, rec in enumerate(records, 2):
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=rec.get(col, ""))
            cell.font = BODY_FONT

        # Color workspace
        ws_cell = ws.cell(row=row_idx, column=workspace_col)
        domain = rec.get("workspace", "")
        if domain in WORKSPACE_FILLS:
            ws_cell.fill = WORKSPACE_FILLS[domain]
            ws_cell.font = WORKSPACE_FONT

        # Color staleness
        tier = rec.get("staleness", "")
        st_cell = ws.cell(row=row_idx, column=staleness_col)
        if tier in STALENESS_FILLS:
            st_cell.fill = STALENESS_FILLS[tier]
            st_cell.font = STALENESS_FONT

    _apply_header(ws)
    _auto_fit(ws)


def _write_naming_conventions(ws) -> None:
    """Tab 3: Naming convention standards for team review."""
    sections = [
        ("Overview", [
            ("Purpose", "Establish consistent naming conventions across all Domo datasets to improve searchability, organization, and team collaboration. These conventions are proposals for team discussion — nothing will be renamed without consensus."),
            ("Structure", "[Environment] - [Domain/System] - [Description] - [Qualifier]"),
            ("Separator", "Always use ' - ' (space-dash-space) as the standard separator between name segments. No underscores, no dots, no inconsistent delimiters."),
        ]),
        ("1. Environment Prefix", [
            ("PROD -", "Production dataset actively used in dashboards and reporting. Optional if there are no dev/test versions of the same dataset."),
            ("DEV -", "Development or work-in-progress dataset not yet promoted to production."),
            ("TEST -", "Test dataset used for validation. Candidate for cleanup after testing is complete."),
            ("DEPRECATED -", "Dataset is no longer in use and is queued for review/deletion. Replaces legacy [DNU] and [Flag for Deletion] tags."),
            ("PATCH -", "Temporary dataset containing one-time data corrections. Should include the date of the patch (YYYY-MM-DD)."),
        ]),
        ("2. Network Names", [
            ("Standard", "Always use full network names: Gilbarco, Speedway, Wayne, Dover. Never abbreviate (no GVR, SPW, etc.)."),
            ("Example", "'Gilbarco - Transactions - Live Hourly' not 'GVR Transactions Live Hourly'"),
        ]),
        ("3. View Prefix", [
            ("View -", "Prefix for derived views of other datasets. Replaces 'View of', 'vw_', and 'Editable DataSet for View of'."),
            ("No nesting", "Never use 'View of View of...' — if a view is two levels deep, the chain should be simplified or the view renamed to reflect its actual purpose."),
            ("Editable View -", "Prefix for editable views only when the editable nature is functionally important."),
        ]),
        ("4. Date Formats", [
            ("ISO standard", "Always use YYYY-MM-DD for specific dates or YYYY-MM for month-level. Never use month names, MMDDYY, or other formats."),
            ("In dataset names", "'PATCH - Gilbarco Transactions - 2025-12-30' not 'Gilbarco Transaction Patch Dec 30'"),
            ("Monthly data", "'All GSTV - SSP Revenue - 2024-08' not 'All GSTV - SSP - August 2024'"),
        ]),
        ("5. File Extensions", [
            ("Remove all", "Never include .xlsx, .csv, .json, .txt in dataset names. These are artifacts of the upload process, not meaningful identifiers."),
            ("Example", "'DXP Conversion Master List - 2025-05-30' not 'DXP Conversion Master List 053025.xlsx'"),
        ]),
        ("6. Legacy Patterns to Replace", [
            ("[DNU]", "Replace with 'DEPRECATED -' prefix"),
            ("[Flag for Deletion]", "Replace with 'DEPRECATED -' prefix"),
            ("[Production]", "Replace with 'PROD -' (no brackets)"),
            ("[PROD]", "Replace with 'PROD -' (no brackets)"),
            ("vw_", "Replace with 'View - '"),
            ("View of View of...", "Simplify to single 'View - '"),
            ("Double spaces", "Collapse to single spaces"),
            ("Underscores in names", "Replace with ' - ' separator (except in technical identifiers)"),
        ]),
        ("7. General Guidelines", [
            ("Be descriptive", "Name should tell someone what the data IS without needing to open it. Avoid cryptic abbreviations."),
            ("Be consistent", "If multiple datasets cover the same concept for different networks, use the same structure: 'Network - Concept - Qualifier'"),
            ("Avoid duplication", "If two datasets have identical names, one of them should be renamed or marked DEPRECATED."),
            ("Keep it scannable", "Frontload the most important information. 'Gilbarco - Transactions - Live Hourly' scans faster than 'Live Hourly Transaction Data for the Gilbarco Network'"),
        ]),
    ]

    row = 1
    for section_title, items in sections:
        # Section header
        ws.cell(row=row, column=1, value=section_title)
        ws.cell(row=row, column=1).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        for c in range(2, 4):
            ws.cell(row=row, column=c).fill = HEADER_FILL
        row += 1

        for term, description in items:
            ws.cell(row=row, column=1, value=term).font = BOLD_FONT
            cell = ws.cell(row=row, column=2, value=description)
            cell.font = BODY_FONT
            cell.alignment = WRAP_ALIGNMENT
            row += 1

        row += 1  # Blank row between sections

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A2"


def _write_proposed_renames(ws, datasets: list[dict[str, Any]]) -> None:
    """Tab 4: Datasets that would be renamed under the new conventions."""
    columns = [
        "dataset_id", "current_name", "proposed_name", "issues_found", "change_count",
    ]

    for col_idx, col in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col)

    records = []
    for ds in datasets:
        name = ds.get("dataset_name", "")
        ds_id = ds.get("dataset_id", "")
        issues = _detect_naming_issues(name)
        proposed = _propose_rename(name)

        if proposed != name and issues:
            records.append({
                "dataset_id": ds_id,
                "current_name": name,
                "proposed_name": proposed,
                "issues_found": " | ".join(i["description"] for i in issues),
                "change_count": len(issues),
            })

    # Sort by issue count descending, then name
    records.sort(key=lambda r: (-r["change_count"], r["current_name"].lower()))

    change_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    for row_idx, rec in enumerate(records, 2):
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=rec.get(col, ""))
            cell.font = BODY_FONT

        # Highlight proposed name if different
        proposed_cell = ws.cell(row=row_idx, column=columns.index("proposed_name") + 1)
        proposed_cell.fill = change_fill

    _apply_header(ws)
    _auto_fit(ws)


def _write_naming_examples(ws, datasets: list[dict[str, Any]]) -> None:
    """Tab 5: Before/after examples grouped by issue type."""
    # Collect examples by issue type
    examples_by_issue: dict[str, list[tuple[str, str]]] = defaultdict(list)

    for ds in datasets:
        name = ds.get("dataset_name", "")
        issues = _detect_naming_issues(name)
        proposed = _propose_rename(name)
        if proposed != name:
            for issue in issues:
                if len(examples_by_issue[issue["description"]]) < 5:  # Max 5 examples per issue
                    examples_by_issue[issue["description"]].append((name, proposed))

    row = 1
    for issue_desc in sorted(examples_by_issue.keys()):
        # Section header
        ws.cell(row=row, column=1, value=issue_desc)
        ws.cell(row=row, column=1).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=2).fill = HEADER_FILL
        ws.cell(row=row, column=3).fill = HEADER_FILL
        row += 1

        ws.cell(row=row, column=1, value="Current Name").font = BOLD_FONT
        ws.cell(row=row, column=2, value="Proposed Name").font = BOLD_FONT
        row += 1

        for current, proposed in examples_by_issue[issue_desc]:
            ws.cell(row=row, column=1, value=current).font = BODY_FONT
            cell = ws.cell(row=row, column=2, value=proposed)
            cell.font = BODY_FONT
            cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            row += 1

        row += 1

    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 80
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def generate_workspace_planner(output_path: str) -> None:
    """Generate the full workspace planning workbook."""
    cache = load_cache()
    if cache is None:
        raise RuntimeError("No cache found. Run a full extraction first.")

    datasets = cache["datasets"]
    lineage = cache["lineage"]
    dataflows = cache["dataflows"]

    # Build analytics
    domain_map = build_domain_map(datasets, lineage)

    wb = Workbook()

    # Tab 1: Workspace Summary
    ws_summary = wb.active
    ws_summary.title = "Workspace Summary"
    _write_workspace_summary(ws_summary, domain_map)

    # Tab 2: Dataset Assignments
    ws_assignments = wb.create_sheet("Dataset Assignments")
    _write_dataset_assignments(ws_assignments, domain_map)

    # Tab 3: Naming Conventions
    ws_naming = wb.create_sheet("Naming Conventions")
    _write_naming_conventions(ws_naming)

    # Tab 4: Proposed Renames
    ws_renames = wb.create_sheet("Proposed Renames")
    _write_proposed_renames(ws_renames, datasets)

    # Tab 5: Naming Examples
    ws_examples = wb.create_sheet("Naming Examples")
    _write_naming_examples(ws_examples, datasets)

    wb.save(output_path)
    logger.info("Workspace planner saved to %s", output_path)


if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.WARNING)

    output_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("output")
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d")
    output_path = output_dir / f"domo_workspace_plan_{timestamp}.xlsx"

    print("Generating workspace planning workbook...", end=" ", flush=True)
    generate_workspace_planner(str(output_path))
    print("✓")
    print(f"Output: {output_path}")
