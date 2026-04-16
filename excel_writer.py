"""Build and format the Excel workbook with all 5 tabs."""

import logging
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from analytics import build_cleanup_candidates, build_dataset_lineage_analysis, build_domain_map

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="3E5170", end_color="3E5170", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)

DATASET_COLUMNS = [
    "dataset_id", "dataset_name", "description", "owner_id", "owner_name",
    "row_count", "column_count", "dataset_type", "pdp_enabled",
    "created_at", "updated_at", "data_current_at",
]

SCHEMA_COLUMNS = [
    "dataset_id", "dataset_name", "column_position", "column_name", "column_type", "definition",
]

DATAFLOW_COLUMNS = [
    "dataflow_id", "dataflow_name", "description", "owner_id", "owner_name",
    "dataflow_type", "status", "input_count", "output_count",
    "last_execution_date", "last_updated_date",
]

LINEAGE_COLUMNS = [
    "dataflow_id", "dataflow_name", "direction", "dataset_id", "dataset_name",
]

DICTIONARY_COLUMNS = [
    "column_name", "column_type", "dataset_count", "commonality",
    "definition", "status", "example_datasets",
]

# Commonality tiers based on how many datasets a column appears in
COMMONALITY_TIERS = [
    (100, "Universal"),
    (50, "Very Common"),
    (20, "Common"),
    (10, "Moderate"),
    (5, "Low"),
    (3, "Niche"),
    (1, "Rare"),
]

COMMONALITY_FILLS = {
    "Universal":    PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid"),  # dark green
    "Very Common":  PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid"),  # green
    "Common":       PatternFill(start_color="689F38", end_color="689F38", fill_type="solid"),  # light green
    "Moderate":     PatternFill(start_color="AFB42B", end_color="AFB42B", fill_type="solid"),  # yellow-green
    "Low":          PatternFill(start_color="FFA000", end_color="FFA000", fill_type="solid"),  # amber
    "Niche":        PatternFill(start_color="E65100", end_color="E65100", fill_type="solid"),  # orange
    "Rare":         PatternFill(start_color="BF360C", end_color="BF360C", fill_type="solid"),  # deep orange
}

COMMONALITY_FONT = Font(name="Arial", size=10, color="FFFFFF")

# --- New analytical tab column lists ---

LINEAGE_ANALYSIS_COLUMNS = [
    "dataset_id", "dataset_name", "role", "feeds_dataflow_count",
    "fed_by_dataflow_count", "downstream_reach", "domain", "department",
    "staleness", "days_since_update", "row_count", "owner_name",
    "feeds_dataflows",
]

CLEANUP_COLUMNS = [
    "type", "id", "name", "domain", "department", "staleness",
    "days_since_update", "has_lineage", "row_count", "owner_name",
    "recommendation", "priority",
]

DOMAIN_MAP_COLUMNS = [
    "domain", "department", "dataset_name", "dataset_id", "role",
    "staleness", "days_since_update", "row_count", "column_count",
    "owner_name", "feeds_dataflow_count", "fed_by_dataflow_count",
]

# Color fills for staleness tiers
STALENESS_FILLS = {
    "Active":      PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid"),  # dark green
    "Stale":       PatternFill(start_color="F9A825", end_color="F9A825", fill_type="solid"),  # yellow
    "Very Stale":  PatternFill(start_color="E65100", end_color="E65100", fill_type="solid"),  # orange
    "Dormant":     PatternFill(start_color="BF360C", end_color="BF360C", fill_type="solid"),  # deep orange
    "Abandoned":   PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid"),  # red
    "Unknown":     PatternFill(start_color="757575", end_color="757575", fill_type="solid"),  # grey
}
STALENESS_FONT = Font(name="Arial", size=10, color="FFFFFF")

ROLE_FILLS = {
    "Source":       PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid"),  # blue
    "Pass-through": PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid"),  # purple
    "Sink":         PatternFill(start_color="00695C", end_color="00695C", fill_type="solid"),  # teal
    "Orphan":       PatternFill(start_color="757575", end_color="757575", fill_type="solid"),  # grey
}
ROLE_FONT = Font(name="Arial", size=10, color="FFFFFF")

RECOMMENDATION_FILLS = {
    "Delete":                          PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid"),
    "Review for Deletion":             PatternFill(start_color="E65100", end_color="E65100", fill_type="solid"),
    "Disable/Delete — Not Executing":  PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid"),
    "Review — Stale but Connected":    PatternFill(start_color="F9A825", end_color="F9A825", fill_type="solid"),
    "Review — Dormant but Connected":  PatternFill(start_color="FF8F00", end_color="FF8F00", fill_type="solid"),
    "Review — Test/Temp Dataset":      PatternFill(start_color="FFA000", end_color="FFA000", fill_type="solid"),
    "Review — Dormant Dataflow":       PatternFill(start_color="FF8F00", end_color="FF8F00", fill_type="solid"),
    "Review — Very Stale Dataflow":    PatternFill(start_color="E65100", end_color="E65100", fill_type="solid"),
}
RECOMMENDATION_FONT_WHITE = Font(name="Arial", size=10, color="FFFFFF")
RECOMMENDATION_FONT_DARK = Font(name="Arial", size=10, color="000000")


def _get_commonality(dataset_count: int) -> str:
    """Return the commonality tier label for a given dataset count."""
    for threshold, label in COMMONALITY_TIERS:
        if dataset_count >= threshold:
            return label
    return "Rare"


def _build_column_dictionary(schemas: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Build a deduplicated column dictionary from schema records.

    Groups by (column_name, column_type), counts datasets, collects example
    dataset names, and assigns a commonality tier.
    """
    from collections import defaultdict

    col_data: dict[tuple[str, str], dict[str, Any]] = {}

    for schema in schemas:
        key = (schema["column_name"], schema["column_type"])
        if key not in col_data:
            col_data[key] = {
                "column_name": schema["column_name"],
                "column_type": schema["column_type"],
                "dataset_count": 0,
                "definition": schema.get("definition", ""),
                "datasets": set(),
            }
        col_data[key]["dataset_count"] += 1
        col_data[key]["datasets"].add(schema.get("dataset_name", ""))
        # Keep the longest definition if multiple exist
        defn = schema.get("definition", "")
        if len(defn) > len(col_data[key]["definition"]):
            col_data[key]["definition"] = defn

    records = []
    for key, data in col_data.items():
        # Pick up to 5 example datasets, sorted alphabetically
        example_ds = sorted(data["datasets"])[:5]
        defined = bool(data["definition"].strip())
        records.append({
            "column_name": data["column_name"],
            "column_type": data["column_type"],
            "dataset_count": data["dataset_count"],
            "commonality": _get_commonality(data["dataset_count"]),
            "definition": data["definition"],
            "status": "Defined" if defined else "Undefined",
            "example_datasets": " | ".join(example_ds),
        })

    # Sort by dataset_count descending, then column_name
    records.sort(key=lambda r: (-r["dataset_count"], r["column_name"].lower(), r["column_type"]))
    return records


def _write_dictionary_sheet(ws, records: list[dict[str, Any]]) -> None:
    """Write the Column Dictionary tab with commonality color coding."""
    columns = DICTIONARY_COLUMNS

    # Header
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data
    commonality_col = columns.index("commonality") + 1
    status_col = columns.index("status") + 1

    for row_idx, record in enumerate(records, 2):
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
            cell.font = BODY_FONT

        # Color-code the commonality cell
        tier = record.get("commonality", "Rare")
        comm_cell = ws.cell(row=row_idx, column=commonality_col)
        if tier in COMMONALITY_FILLS:
            comm_cell.fill = COMMONALITY_FILLS[tier]
            comm_cell.font = COMMONALITY_FONT

        # Highlight undefined columns
        if record.get("status") == "Undefined":
            status_cell = ws.cell(row=row_idx, column=status_col)
            status_cell.font = Font(name="Arial", size=10, color="CC0000")

    _apply_header_formatting(ws)
    _auto_fit_columns(ws)


def _apply_header_formatting(ws) -> None:
    """Apply navy header with white bold text and freeze top row."""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"


def _auto_fit_columns(ws) -> None:
    """Auto-fit column widths based on content, with a max width cap."""
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), 1):
        max_len = 0
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        adjusted = min(max_len + 2, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def _write_sheet(ws, columns: list[str], records: list[dict[str, Any]]) -> None:
    """Write header row and data rows, then format."""
    # Header
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data
    for row_idx, record in enumerate(records, 2):
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
            cell.font = BODY_FONT

    _apply_header_formatting(ws)
    _auto_fit_columns(ws)


def _write_lineage_analysis_sheet(ws, records: list[dict[str, Any]]) -> None:
    """Write the Dataset Lineage Analysis tab with color-coded roles and staleness."""
    columns = LINEAGE_ANALYSIS_COLUMNS

    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    role_col = columns.index("role") + 1
    staleness_col = columns.index("staleness") + 1

    for row_idx, record in enumerate(records, 2):
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
            cell.font = BODY_FONT

        # Color-code role
        role = record.get("role", "")
        role_cell = ws.cell(row=row_idx, column=role_col)
        if role in ROLE_FILLS:
            role_cell.fill = ROLE_FILLS[role]
            role_cell.font = ROLE_FONT

        # Color-code staleness
        tier = record.get("staleness", "")
        stale_cell = ws.cell(row=row_idx, column=staleness_col)
        if tier in STALENESS_FILLS:
            stale_cell.fill = STALENESS_FILLS[tier]
            stale_cell.font = STALENESS_FONT

    _apply_header_formatting(ws)
    _auto_fit_columns(ws)


def _write_cleanup_sheet(ws, records: list[dict[str, Any]]) -> None:
    """Write the Cleanup Candidates tab with color-coded recommendations."""
    columns = CLEANUP_COLUMNS

    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    staleness_col = columns.index("staleness") + 1
    rec_col = columns.index("recommendation") + 1

    for row_idx, record in enumerate(records, 2):
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
            cell.font = BODY_FONT

        # Color-code staleness
        tier = record.get("staleness", "")
        stale_cell = ws.cell(row=row_idx, column=staleness_col)
        if tier in STALENESS_FILLS:
            stale_cell.fill = STALENESS_FILLS[tier]
            stale_cell.font = STALENESS_FONT

        # Color-code recommendation
        rec = record.get("recommendation", "")
        rec_cell = ws.cell(row=row_idx, column=rec_col)
        if rec in RECOMMENDATION_FILLS:
            rec_cell.fill = RECOMMENDATION_FILLS[rec]
            # Use dark text on yellow-ish backgrounds
            if rec in ("Review — Stale but Connected", "Review — Test/Temp Dataset"):
                rec_cell.font = RECOMMENDATION_FONT_DARK
            else:
                rec_cell.font = RECOMMENDATION_FONT_WHITE

    _apply_header_formatting(ws)
    _auto_fit_columns(ws)


def _write_domain_map_sheet(ws, records: list[dict[str, Any]]) -> None:
    """Write the Domain Map tab with color-coded roles and staleness."""
    columns = DOMAIN_MAP_COLUMNS

    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    role_col = columns.index("role") + 1
    staleness_col = columns.index("staleness") + 1

    for row_idx, record in enumerate(records, 2):
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
            cell.font = BODY_FONT

        # Color-code role
        role = record.get("role", "")
        role_cell = ws.cell(row=row_idx, column=role_col)
        if role in ROLE_FILLS:
            role_cell.fill = ROLE_FILLS[role]
            role_cell.font = ROLE_FONT

        # Color-code staleness
        tier = record.get("staleness", "")
        stale_cell = ws.cell(row=row_idx, column=staleness_col)
        if tier in STALENESS_FILLS:
            stale_cell.fill = STALENESS_FILLS[tier]
            stale_cell.font = STALENESS_FONT

    _apply_header_formatting(ws)
    _auto_fit_columns(ws)


def _write_glossary_sheet(ws) -> None:
    """Write the Glossary tab explaining terms used across analytical tabs."""
    sections = [
        ("Dataset Roles (Lineage Analysis, Domain Map)", [
            ("Source", "A dataset that feeds into one or more dataflows but is not itself produced by any dataflow. These are your foundational datasets — connectors, uploads, and API imports that everything else depends on."),
            ("Sink", "A dataset that is produced by a dataflow but does not feed into any other dataflow. These are end-of-line outputs, typically powering cards and dashboards."),
            ("Pass-through", "A dataset that is both produced by a dataflow AND feeds into another dataflow. These are intermediate transformation steps in a data pipeline."),
            ("Orphan", "A dataset with no lineage connections at all — it neither feeds into nor is produced by any dataflow. May be a standalone connector, a manual upload, or a candidate for cleanup."),
        ]),
        ("Staleness Tiers", [
            ("Active (0-7 days)", "Data was updated within the last week. The dataset is healthy and current."),
            ("Stale (8-30 days)", "Data is between 1-4 weeks old. May be on a monthly refresh cycle, or may need attention."),
            ("Very Stale (31-90 days)", "Data is 1-3 months old. Likely not being actively maintained unless on a quarterly schedule."),
            ("Dormant (91-365 days)", "Data is 3-12 months old. Strong candidate for review — may be abandoned or obsolete."),
            ("Abandoned (>365 days)", "Data has not been updated in over a year. High-priority cleanup candidate unless intentionally archived."),
        ]),
        ("Cleanup Recommendations", [
            ("Delete", "Dataset is abandoned AND has no lineage connections. Safe to remove after owner confirmation."),
            ("Review for Deletion", "Dataset is very stale with no lineage connections. Likely safe to remove but warrants a quick check."),
            ("Disable/Delete — Not Executing", "Dataflow has not executed in over a year. Should be disabled or removed."),
            ("Review — Stale but Connected", "Dataset is abandoned but still has lineage connections. Removing it could break downstream pipelines — investigate before acting."),
            ("Review — Dormant but Connected", "Dataset is dormant but still connected in lineage. May be on an infrequent schedule or may need intervention."),
            ("Review — Test/Temp Dataset", "Dataset name suggests it is a test, temporary, sandbox, or archived dataset. Review for cleanup regardless of staleness."),
            ("Review — Dormant Dataflow", "Dataflow has not executed in 3-12 months. May need to be re-enabled or removed."),
            ("Review — Very Stale Dataflow", "Dataflow has not executed in 1-3 months. Check if it should still be running."),
        ]),
        ("Lineage Metrics", [
            ("Downstream Reach", "The total number of unique datasets that can be reached by following the output chains from a given dataset through dataflows. A high number means many other datasets depend on this one — it is foundational."),
            ("Feeds Dataflow Count", "How many dataflows use this dataset as an input."),
            ("Fed By Dataflow Count", "How many dataflows produce this dataset as an output."),
        ]),
        ("Column Dictionary Terms", [
            ("Commonality — Universal (100+)", "Column appears in 100 or more datasets. Core system-wide field."),
            ("Commonality — Very Common (50-99)", "Column appears in 50-99 datasets. Widely used across domains."),
            ("Commonality — Common (20-49)", "Column appears in 20-49 datasets. Shared across multiple data areas."),
            ("Commonality — Moderate (10-19)", "Column appears in 10-19 datasets. Used in a specific set of related datasets."),
            ("Commonality — Low (5-9)", "Column appears in 5-9 datasets. Limited usage."),
            ("Commonality — Niche (3-4)", "Column appears in 3-4 datasets. Specialized usage."),
            ("Commonality — Rare (1-2)", "Column appears in only 1-2 datasets. Unique or dataset-specific field."),
            ("Defined / Undefined", "Whether a human-written definition exists for this column in the data dictionary."),
        ]),
        ("Domain Classifications (maps to Domain Workspaces)", [
            ("Sites & Locations", "Master station list, site configuration, GBase, MongoDB site data, TDLinx, DXP/DXPromote conversion data, site status history, market plan site lists, DMA GTVID creation — owned by Data Engineering / Network Operations."),
            ("Proof of Play", "Gilbarco ICS, Applause, Dover DXPromote POP data, terminal play data, POP monitoring, sites reporting POP — owned by Network Operations."),
            ("Impressions", "Network Validated Impressions (NVI), Campaign Validated Impressions (CVI), hourly impression multipliers, Comscore impression data — owned by Data & Analytics."),
            ("Transactions", "Live hourly transaction data by network (Gilbarco, Speedway, Wayne, Dover), transaction monitoring, patches, unvalidated transactions — owned by Data & Analytics."),
            ("Revenue & Monetization", "Revenue share invoices, clawback calculations, retailer campaign revenue, CPM floors, Pluto revenue projections, programmatic revenue (as a subsection) — owned by Finance."),
            ("Programmatic Operations", "Vistar diagnostics, venue availability, fill rates, auction/bid data, SSP site inventory, PlaceExchange inventory/PMP pacing, sell-through analysis — owned by Programmatic."),
            ("Programmatic", "Vistar exchange revenue, PlaceExchange revenue, Broadsign/Hivestack/Magnite SSP revenue, general programmatic performance and tracking — owned by Programmatic. Revenue datasets here are candidates for the Revenue & Monetization subsection."),
            ("RPA", "Rotation and play assignment — RPA conversion, scheduling, utilization, email loads, Nuxeo asset checks, RPA base data sets — owned by Ad Operations."),
            ("Traffic Instructions", "Traffic instruction datasets — encompasses both RPA and programmatic delivery instructions — owned by Ad Operations."),
            ("Managed Services", "Retailer-specific operational datasets for managed service partners — Casey's, Speedway, Circle K, Kwik Trip, etc. Includes schedules, contracts, deployment trackers, site lists — owned by Ad Operations."),
            ("Campaigns & Delivery", "Campaign delivery reporting, campaign tracking, IO lists, campaign launch monitoring, advertiser-specific campaign datasets (State Farm, Fairlife, etc.) — owned by Ad Operations."),
            ("Salesforce / CRM", "Salesforce connector data — opportunities, accounts, contacts, SF integration and reconciliation datasets — owned by Sales."),
            ("Monitoring & Governance", "DomoStats governance datasets, PMAR, data quality checks, alerting, general monitoring and diagnostics — owned by Data Engineering."),
            ("Site Analytics", "Site scoring, platinum classification, lookalike models, site metrics — owned by Data & Analytics."),
            ("Engineering", "Jira, sprint, worklog, and agile project data — owned by Engineering."),
            ("Test / Temp / Archive", "Test, temporary, sandbox, deprecated, or archived datasets — flagged as Cleanup Candidates."),
            ("Other / Unclassified", "Datasets that did not match any domain classification pattern. Needs manual review for workspace assignment."),
        ]),
    ]

    row = 1
    for section_title, items in sections:
        # Section header
        ws.cell(row=row, column=1, value=section_title)
        ws.cell(row=row, column=1).font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=2).fill = HEADER_FILL
        ws.cell(row=row, column=3).fill = HEADER_FILL
        row += 1

        # Column sub-headers
        ws.cell(row=row, column=1, value="Term").font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=row, column=2, value="Definition").font = Font(name="Arial", size=10, bold=True)
        row += 1

        for term, definition in items:
            ws.cell(row=row, column=1, value=term).font = BODY_FONT
            cell = ws.cell(row=row, column=2, value=definition)
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

        row += 1  # Blank row between sections

    # Set column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 100
    ws.freeze_panes = "A2"


def _write_extraction_log(
    ws,
    extraction_time: datetime,
    dataset_count: int,
    schema_count: int,
    dictionary_count: int,
    dataflow_count: int,
    lineage_count: int,
    lineage_analysis_count: int,
    cleanup_count: int,
    domain_map_count: int,
    errors: list[dict[str, str]],
) -> None:
    """Write the Extraction Log tab with run metadata."""
    ws.cell(row=1, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=1, column=2).fill = HEADER_FILL
    ws.freeze_panes = "A2"

    rows = [
        ("Extraction Date/Time", extraction_time.strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Datasets", dataset_count),
        ("Total Schema Columns", schema_count),
        ("Unique Columns (Dictionary)", dictionary_count),
        ("Total Dataflows", dataflow_count),
        ("Total Lineage Relationships", lineage_count),
        ("Lineage Analysis Entries", lineage_analysis_count),
        ("Cleanup Candidates", cleanup_count),
        ("Domain Map Entries", domain_map_count),
        ("Items Skipped Due to Errors", len(errors)),
    ]

    for row_idx, (metric, value) in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=metric).font = BODY_FONT
        ws.cell(row=row_idx, column=2, value=value).font = BODY_FONT

    if errors:
        # Add error detail section
        error_start = len(rows) + 3
        ws.cell(row=error_start, column=1, value="Skipped Items").font = Font(name="Arial", size=10, bold=True)

        error_headers = ["Type", "ID", "Name", "Error"]
        for col_idx, h in enumerate(error_headers, 1):
            cell = ws.cell(row=error_start + 1, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        for err_idx, err in enumerate(errors, error_start + 2):
            ws.cell(row=err_idx, column=1, value=err.get("type", "")).font = BODY_FONT
            ws.cell(row=err_idx, column=2, value=err.get("id", "")).font = BODY_FONT
            ws.cell(row=err_idx, column=3, value=err.get("name", "")).font = BODY_FONT
            ws.cell(row=err_idx, column=4, value=err.get("error", "")).font = BODY_FONT

    _auto_fit_columns(ws)


def write_workbook(
    output_path: str,
    datasets: list[dict[str, Any]],
    schemas: list[dict[str, Any]],
    dataflows: list[dict[str, Any]],
    lineage: list[dict[str, Any]],
    errors: list[dict[str, str]],
    extraction_time: datetime,
) -> None:
    """Create the full Excel workbook with all 9 tabs."""
    wb = Workbook()

    # Tab 1: Datasets
    ws_datasets = wb.active
    ws_datasets.title = "Datasets"
    _write_sheet(ws_datasets, DATASET_COLUMNS, datasets)

    # Tab 2: Dataset Schemas
    ws_schemas = wb.create_sheet("Dataset Schemas")
    _write_sheet(ws_schemas, SCHEMA_COLUMNS, schemas)

    # Tab 3: Column Dictionary (deduplicated, sorted by column name)
    ws_dict = wb.create_sheet("Column Dictionary")
    dictionary_records = _build_column_dictionary(schemas)
    _write_dictionary_sheet(ws_dict, dictionary_records)

    # Tab 4: Dataflows
    ws_dataflows = wb.create_sheet("Dataflows")
    _write_sheet(ws_dataflows, DATAFLOW_COLUMNS, dataflows)

    # Tab 5: Dataflow Lineage
    ws_lineage = wb.create_sheet("Dataflow Lineage")
    _write_sheet(ws_lineage, LINEAGE_COLUMNS, lineage)

    # Tab 6: Dataset Lineage Analysis
    ws_lineage_analysis = wb.create_sheet("Lineage Analysis")
    lineage_analysis_records = build_dataset_lineage_analysis(datasets, lineage, dataflows)
    _write_lineage_analysis_sheet(ws_lineage_analysis, lineage_analysis_records)

    # Tab 7: Cleanup Candidates
    ws_cleanup = wb.create_sheet("Cleanup Candidates")
    cleanup_records = build_cleanup_candidates(datasets, lineage, dataflows)
    _write_cleanup_sheet(ws_cleanup, cleanup_records)

    # Tab 8: Domain Map
    ws_domain = wb.create_sheet("Domain Map")
    domain_map_records = build_domain_map(datasets, lineage)
    _write_domain_map_sheet(ws_domain, domain_map_records)

    # Tab 9: Glossary
    ws_glossary = wb.create_sheet("Glossary")
    _write_glossary_sheet(ws_glossary)

    # Tab 10: Extraction Log (always last)
    ws_log = wb.create_sheet("Extraction Log")
    _write_extraction_log(
        ws_log,
        extraction_time=extraction_time,
        dataset_count=len(datasets),
        schema_count=len(schemas),
        dictionary_count=len(dictionary_records),
        dataflow_count=len(dataflows),
        lineage_count=len(lineage),
        lineage_analysis_count=len(lineage_analysis_records),
        cleanup_count=len(cleanup_records),
        domain_map_count=len(domain_map_records),
        errors=errors,
    )

    wb.save(output_path)
    logger.info("Workbook saved to %s", output_path)
