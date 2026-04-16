#!/usr/bin/env python3
"""Remove datasets marked for removal after the cleanup deadline.

Reads returned cleanup spreadsheets and deletes datasets where:
  - Owner marked "Remove" in the Decision column, OR
  - No response was received by the deadline (configurable)

Creates a rollback manifest with all deleted dataset metadata before deletion.

Usage:
    python3 apply_removals.py --dry-run                    # Preview (default)
    python3 apply_removals.py --execute                    # Delete with confirmation
    python3 apply_removals.py --execute --include-no-response  # Also delete non-responses
"""

import argparse
import csv
import json
import logging
import os
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from domo_client import DomoClient

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

OUTPUT_DIR = Path(__file__).resolve().parent / "output"
ROLLOUT_DIR = OUTPUT_DIR / "owner_rollouts"
LOG_DIR = OUTPUT_DIR / "automation_logs"
CACHE_PATH = Path(__file__).resolve().parent / ".cache" / "latest.json"


def load_removals_from_spreadsheets(include_no_response: bool = False) -> list[dict]:
    """Load items marked for removal from returned cleanup spreadsheets."""
    removals = []

    if not ROLLOUT_DIR.exists():
        logger.error("Rollout directory not found: %s", ROLLOUT_DIR)
        return removals

    for xlsx_path in sorted(ROLLOUT_DIR.glob("cleanup_review_*.xlsx")):
        logger.info("Reading %s", xlsx_path.name)
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)

        for sheet_name in ("Datasets", "Dataflows"):
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            headers = [str(h).strip() if h else "" for h in rows[0]]
            col_map = {h.lower(): i for i, h in enumerate(headers)}

            id_col = col_map.get("id") or col_map.get("dataset id") or col_map.get("dataflow id")
            name_col = col_map.get("name") or col_map.get("dataset name") or col_map.get("dataflow name")
            decision_col = col_map.get("decision")
            owner_col = col_map.get("owner")
            cards_col = col_map.get("cards affected")

            if id_col is None or name_col is None:
                continue

            for row in rows[1:]:
                if not row or not row[id_col]:
                    continue

                item_id = str(row[id_col]).strip()
                current_name = str(row[name_col]).strip() if row[name_col] else ""
                decision = str(row[decision_col]).strip().lower() if decision_col is not None and row[decision_col] else ""
                owner = str(row[owner_col]).strip() if owner_col is not None and row[owner_col] else ""
                cards = int(row[cards_col]) if cards_col is not None and row[cards_col] else 0
                item_type = "dataset" if sheet_name == "Datasets" else "dataflow"

                should_remove = False
                reason = ""

                if decision == "remove":
                    should_remove = True
                    reason = "Marked 'Remove' by owner"
                elif include_no_response and not decision:
                    should_remove = True
                    reason = "No response by deadline"

                if should_remove:
                    removals.append({
                        "id": item_id,
                        "name": current_name,
                        "type": item_type,
                        "owner": owner,
                        "cards_affected": cards,
                        "reason": reason,
                        "source": xlsx_path.name,
                    })

        wb.close()

    return removals


def save_rollback_manifest(removals: list[dict], cache: dict) -> Path:
    """Save a rollback manifest with full dataset metadata before deletion."""
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    manifest_path = LOG_DIR / f"removal_rollback_manifest_{timestamp}.json"

    ds_lookup = {ds["dataset_id"]: ds for ds in cache.get("datasets", [])}
    df_lookup = {df["dataflow_id"]: df for df in cache.get("dataflows", [])}

    manifest = {
        "created_at": datetime.now().isoformat(),
        "total_removals": len(removals),
        "items": [],
    }

    for r in removals:
        item_data = {
            "id": r["id"],
            "name": r["name"],
            "type": r["type"],
            "owner": r["owner"],
            "reason": r["reason"],
            "cards_affected": r["cards_affected"],
        }

        # Attach full metadata from cache for rollback reference
        if r["type"] == "dataset" and r["id"] in ds_lookup:
            item_data["cached_metadata"] = ds_lookup[r["id"]]
        elif r["type"] == "dataflow" and r["id"] in df_lookup:
            item_data["cached_metadata"] = df_lookup[r["id"]]

        manifest["items"].append(item_data)

    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, default=str)

    logger.info("Rollback manifest saved to: %s", manifest_path)
    return manifest_path


def apply_removals(removals: list[dict], dry_run: bool = True) -> dict:
    """Delete datasets via the Domo API."""
    # Separate datasets from dataflows
    dataset_removals = [r for r in removals if r["type"] == "dataset"]
    dataflow_removals = [r for r in removals if r["type"] == "dataflow"]

    if dataflow_removals:
        logger.warning(
            "Skipping %d dataflow removals — dataflows cannot be deleted via API",
            len(dataflow_removals),
        )

    # Warn about items with dashboard impact
    impacted = [r for r in dataset_removals if r["cards_affected"] > 0]
    if impacted:
        logger.warning(
            "\n⚠️  WARNING: %d datasets power live dashboard cards!",
            len(impacted),
        )
        for r in impacted:
            logger.warning("  %s — %d cards affected", r["name"][:50], r["cards_affected"])

    stats = {
        "total": len(dataset_removals),
        "success": 0,
        "failed": 0,
        "skipped_dataflows": len(dataflow_removals),
        "with_dashboard_impact": len(impacted),
        "dry_run": dry_run,
    }

    if not dataset_removals:
        logger.info("No datasets to remove.")
        return stats

    # Save rollback manifest
    with open(CACHE_PATH, encoding="utf-8") as f:
        cache = json.load(f)
    save_rollback_manifest(removals, cache)

    LOG_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = LOG_DIR / f"removals_{timestamp}.csv"

    if dry_run:
        logger.info("\n" + "=" * 70)
        logger.info("DRY RUN — No datasets will be deleted")
        logger.info("=" * 70)
        logger.info("Would delete %d datasets:\n", len(dataset_removals))

        with open(log_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["dataset_id", "name", "owner", "reason", "cards_affected", "status"])
            for r in dataset_removals:
                flag = " ⚠️" if r["cards_affected"] > 0 else ""
                logger.info("  [%s] %s — %s%s", r["reason"], r["name"][:50], r["owner"], flag)
                writer.writerow([r["id"], r["name"], r["owner"], r["reason"], r["cards_affected"], "dry_run"])

        logger.info("\nDry run log saved to: %s", log_path)
        return stats

    # Confirmation prompt
    print(f"\n{'=' * 70}")
    print(f"⚠️  ABOUT TO DELETE {len(dataset_removals)} DATASETS")
    if impacted:
        print(f"⚠️  {len(impacted)} of these power live dashboard cards!")
    print(f"{'=' * 70}")
    confirm = input("\nType 'DELETE' to confirm, anything else to abort: ")
    if confirm.strip() != "DELETE":
        logger.info("Aborted by user.")
        stats["dry_run"] = True
        return stats

    # Live execution
    client = DomoClient(
        client_id=os.environ["DOMO_CLIENT_ID"],
        client_secret=os.environ["DOMO_CLIENT_SECRET"],
    )
    client.authenticate()
    logger.info("Authenticated with Domo API")

    results = []
    for i, r in enumerate(dataset_removals, 1):
        logger.info("[%d/%d] Deleting %s (%s)", i, len(dataset_removals), r["name"][:50], r["reason"])

        success = client.delete_dataset(r["id"])
        if success:
            stats["success"] += 1
            results.append([r["id"], r["name"], r["owner"], r["reason"], r["cards_affected"], "deleted"])
            logger.info("  ✓ Deleted")
        else:
            stats["failed"] += 1
            results.append([r["id"], r["name"], r["owner"], r["reason"], r["cards_affected"], "failed"])
            logger.error("  ✗ Failed")

    with open(log_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["dataset_id", "name", "owner", "reason", "cards_affected", "status"])
        writer.writerows(results)

    logger.info("\nResults log saved to: %s", log_path)
    return stats


def main():
    parser = argparse.ArgumentParser(description="Remove datasets marked for deletion in Domo")
    parser.add_argument("--dry-run", action="store_true", default=True,
                        help="Preview deletions without executing (default: True)")
    parser.add_argument("--execute", action="store_true",
                        help="Actually delete datasets (overrides --dry-run)")
    parser.add_argument("--include-no-response", action="store_true",
                        help="Also delete items with no response (deadline policy)")
    args = parser.parse_args()

    dry_run = not args.execute

    removals = load_removals_from_spreadsheets(
        include_no_response=args.include_no_response,
    )
    logger.info("Found %d items marked for removal", len(removals))

    if not removals:
        logger.info("Nothing to remove. Exiting.")
        return

    stats = apply_removals(removals, dry_run=dry_run)

    print(f"\n{'=' * 60}")
    print("REMOVAL SUMMARY")
    print(f"{'=' * 60}")
    print(f"  Mode:               {'DRY RUN' if stats['dry_run'] else 'LIVE'}")
    print(f"  Datasets to remove: {stats['total']}")
    print(f"  With dashboard impact: {stats['with_dashboard_impact']}")
    if not dry_run and not stats["dry_run"]:
        print(f"  Deleted:            {stats['success']}")
        print(f"  Failed:             {stats['failed']}")
    print(f"  Dataflows skipped:  {stats['skipped_dataflows']} (requires UI)")


if __name__ == "__main__":
    main()
