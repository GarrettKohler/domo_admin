"""Generate the team runbook and interview worksheets.

Adds tabs to the workspace planning workbook:
- Team Runbook: step-by-step workflow for the cleanup/organization project
- Unclassified Review: interview sheet for triaging unclassified datasets
- Owner Action Map: interview sheet for assigning cleanup owners
"""

import csv
import json
import logging
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from analytics import _classify_domain, build_cleanup_candidates, build_domain_map
from extractors import load_cache

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="3E5170", end_color="3E5170", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)
SECTION_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")

PHASE_FILLS = {
    "Phase 1": PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid"),
    "Phase 2": PatternFill(start_color="00695C", end_color="00695C", fill_type="solid"),
    "Phase 3": PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid"),
    "Phase 4": PatternFill(start_color="E65100", end_color="E65100", fill_type="solid"),
    "Phase 5": PatternFill(start_color="AD1457", end_color="AD1457", fill_type="solid"),
}
PHASE_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")

ACTION_FILLS = {
    "Delete":  PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid"),
    "Review":  PatternFill(start_color="F9A825", end_color="F9A825", fill_type="solid"),
    "Keep":    PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid"),
    "Rename":  PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid"),
    "Assign":  PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid"),
}


def _write_runbook(ws) -> None:
    """Write the team runbook with phased workflow."""
    phases = [
        ("Phase 1: Triage & Quick Wins (Week 1, Days 1-2)", "Goal: Eliminate obvious cleanup items and establish ownership.", [
            ("1.1", "Review Test/Temp/Archive datasets", "Filter the Cleanup Candidates tab for 'Test / Temp / Archive' domain. These 218 datasets are the easiest wins.", "All", "Delete confirmed test/temp datasets. Move uncertain ones to Phase 2."),
            ("1.2", "Review exact duplicate datasets", "Use the Duplicate Analysis tab. For each group of identical-name datasets, determine which is active and which are orphaned copies.", "All", "Flag duplicates for deletion, keeping only the active version."),
            ("1.3", "Classify unclassified datasets", "Each team member reviews their section of the Unclassified Review tab. Assign each dataset to a workspace or mark it for deletion.", "All team leads", "Every dataset has a workspace assignment or deletion flag."),
            ("1.4", "Assign cleanup owners", "Team leads fill in the Owner Action Map tab for their domain. Each cleanup candidate gets an assigned reviewer.", "Team leads", "Every cleanup candidate has an owner."),
        ]),
        ("Phase 2: Cleanup Execution (Week 1, Days 3-5)", "Goal: Execute deletions and begin workspace organization.", [
            ("2.1", "Delete confirmed items", "Work through your assigned deletion list in the Owner Action Map. Verify each dataset has no active cards or dependencies before deleting in the Domo UI.", "Assigned owners", "All 'Delete' items either deleted or escalated."),
            ("2.2", "Disable abandoned dataflows", "Filter Cleanup Candidates for 'Dataflow' type with 'Disable/Delete' recommendation. Disable in the Domo UI.", "Data Engineering", "Abandoned dataflows disabled."),
            ("2.3", "Review 'Stale but Connected' items", "These datasets are old but still have lineage connections. Trace the lineage to determine if downstream consumers are also stale.", "Data Engineering", "Decision made: delete chain, refresh, or keep as-is."),
        ]),
        ("Phase 3: Workspace Setup (Week 2, Days 1-2)", "Goal: Create workspace structure and begin dataset assignment.", [
            ("3.1", "Create workspaces in Domo", "Create the 15 workspaces from the Workspace Summary tab (excluding Test/Temp/Archive and Other/Unclassified).", "Domo Admin", "All workspaces created with correct names."),
            ("3.2", "Create subsections", "Within each workspace, create subpages matching the subsections shown in the Workspace Summary tab.", "Domo Admin", "Subsection structure in place."),
            ("3.3", "Assign datasets to workspaces", "Use the Dataset Assignments tab as a checklist. Filter by workspace and work through each one.", "Team leads (by domain)", "All active datasets assigned to a workspace."),
        ]),
        ("Phase 4: Naming & Definitions (Week 2, Days 3-5)", "Goal: Standardize naming and complete the data dictionary.", [
            ("4.1", "Review proposed renames", "Review the Proposed Renames tab as a team. Approve, modify, or reject each proposed rename.", "All", "Team consensus on naming standards and specific renames."),
            ("4.2", "Execute approved renames", "Apply naming changes in the Domo UI (or via API script if volume warrants it).", "Data Engineering", "Naming conventions applied."),
            ("4.3", "Complete column definitions", "Each team fills in definitions for their domain using the interview tool or the exported CSV. Focus on high-impact columns (appearing in 5+ datasets) first.", "All team leads", "Definition coverage above 90% for high-impact columns."),
        ]),
        ("Phase 5: Validation & Handoff (End of Week 2)", "Goal: Verify everything is clean and documented.", [
            ("5.1", "Rebuild inventory workbook", "Run `python3 main.py --rebuild` to regenerate the workbook with all updates.", "Data Engineering", "Fresh workbook reflecting all changes."),
            ("5.2", "Validate workspace assignments", "Spot-check that datasets are in the correct workspace and subsection.", "Team leads", "No misplaced datasets."),
            ("5.3", "Review definition coverage", "Run `python3 interview.py --stats` to verify coverage targets are met.", "Data Engineering", "Coverage targets met."),
            ("5.4", "Archive cleanup records", "Save the final cleanup workbook as a record of what was deleted, renamed, and reorganized.", "Project owner", "Audit trail complete."),
        ]),
    ]

    row = 1
    # Title
    ws.cell(row=row, column=1, value="Domo Cleanup & Organization Runbook")
    ws.cell(row=row, column=1).font = Font(name="Arial", size=14, bold=True)
    row += 1
    ws.cell(row=row, column=1, value=f"Generated {datetime.now().strftime('%Y-%m-%d')} — Review and adapt to your team's capacity")
    ws.cell(row=row, column=1).font = Font(name="Arial", size=10, italic=True, color="666666")
    row += 2

    columns = ["Step", "Task", "Details", "Owner", "Done Criteria"]
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1

    for phase_title, phase_goal, steps in phases:
        # Phase header
        phase_key = phase_title.split(":")[0].strip()
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row, column=col_idx)
            if col_idx == 1:
                cell.value = phase_title
            elif col_idx == 2:
                cell.value = phase_goal
            cell.font = PHASE_FONT
            cell.fill = PHASE_FILLS.get(phase_key, HEADER_FILL)
        row += 1

        for step_num, task, details, owner, done in steps:
            ws.cell(row=row, column=1, value=step_num).font = BOLD_FONT
            ws.cell(row=row, column=2, value=task).font = BOLD_FONT
            cell = ws.cell(row=row, column=3, value=details)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            ws.cell(row=row, column=4, value=owner).font = BODY_FONT
            cell = ws.cell(row=row, column=5, value=done)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            row += 1

        row += 1  # Gap between phases

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 45
    ws.freeze_panes = "A5"


def _write_unclassified_review(ws, datasets: list[dict], lineage: list[dict]) -> None:
    """Write the Unclassified Review interview sheet."""
    # Build lineage context
    ds_feeds = defaultdict(set)
    ds_fed_by = defaultdict(set)
    for rec in lineage:
        ds_id = rec.get("dataset_id", "")
        df_id = rec.get("dataflow_id", "")
        df_name = rec.get("dataflow_name", "")
        direction = rec.get("direction", "")
        if direction == "Input":
            ds_feeds[ds_id].add(df_name or df_id)
        elif direction == "Output":
            ds_fed_by[ds_id].add(df_name or df_id)

    # Find unclassified
    unclassified = []
    for ds in datasets:
        domain, dept = _classify_domain(ds.get("dataset_name", ""))
        if domain == "Other / Unclassified":
            ds_id = ds.get("dataset_id", "")
            feeds = sorted(ds_feeds.get(ds_id, set()))[:3]
            fed_by = sorted(ds_fed_by.get(ds_id, set()))[:3]
            unclassified.append({
                "dataset_name": ds.get("dataset_name", ""),
                "dataset_id": ds_id,
                "owner_name": ds.get("owner_name", ""),
                "row_count": ds.get("row_count", 0),
                "data_current_at": ds.get("data_current_at", ""),
                "feeds_into": " | ".join(feeds) if feeds else "(none)",
                "produced_by": " | ".join(fed_by) if fed_by else "(none)",
                "assign_to_workspace": "",  # Blank for team to fill
                "action": "",  # Blank: Keep / Delete / Review
                "notes": "",  # Blank for team notes
            })

    unclassified.sort(key=lambda r: r["dataset_name"].lower())

    columns = [
        "dataset_name", "dataset_id", "owner_name", "row_count",
        "data_current_at", "feeds_into", "produced_by",
        "assign_to_workspace", "action", "notes",
    ]

    # Instructions row
    instruction_font = Font(name="Arial", size=10, bold=True, italic=True, color="1565C0")
    ws.cell(row=1, column=1, value="INSTRUCTIONS: For each unclassified dataset, fill in 'assign_to_workspace' with the workspace name from the Workspace Summary tab, and 'action' with Keep/Delete/Review.")
    ws.cell(row=1, column=1).font = instruction_font
    for c in range(2, len(columns) + 1):
        ws.cell(row=1, column=c).font = instruction_font

    # Headers
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Highlight the fillable columns
    fill_cols = [columns.index("assign_to_workspace") + 1, columns.index("action") + 1, columns.index("notes") + 1]
    input_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    for row_idx, rec in enumerate(unclassified, 3):
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=rec.get(col, ""))
            cell.font = BODY_FONT
            if col_idx in fill_cols:
                cell.fill = input_fill

    # Auto-fit
    col_widths = {"dataset_name": 60, "dataset_id": 38, "owner_name": 20, "row_count": 12,
                  "data_current_at": 22, "feeds_into": 50, "produced_by": 50,
                  "assign_to_workspace": 25, "action": 12, "notes": 40}
    for col_idx, col in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col, 15)

    ws.freeze_panes = "A3"


def _write_owner_action_map(ws, datasets: list[dict], lineage: list[dict], dataflows: list[dict]) -> None:
    """Write the Owner Action Map interview sheet for cleanup assignment."""
    cleanup = build_cleanup_candidates(datasets, lineage, dataflows)

    # Enrich with workspace
    for rec in cleanup:
        domain, dept = _classify_domain(rec.get("name", ""))
        rec["workspace"] = domain
        rec["department"] = dept
        rec["assigned_to"] = ""  # Blank for team to fill
        rec["action_taken"] = ""  # Blank: Deleted / Kept / Renamed / Deferred
        rec["action_date"] = ""  # Blank
        rec["notes"] = ""  # Blank

    columns = [
        "workspace", "department", "type", "name", "id",
        "staleness", "days_since_update", "has_lineage", "recommendation",
        "assigned_to", "action_taken", "action_date", "notes",
    ]

    # Instructions
    instruction_font = Font(name="Arial", size=10, bold=True, italic=True, color="1565C0")
    ws.cell(row=1, column=1, value="INSTRUCTIONS: Team leads assign themselves or team members in 'assigned_to'. After review, fill in 'action_taken' (Deleted/Kept/Renamed/Deferred) and 'action_date'.")
    ws.cell(row=1, column=1).font = instruction_font

    # Headers
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Fillable columns
    fill_col_names = {"assigned_to", "action_taken", "action_date", "notes"}
    fill_col_indices = {columns.index(c) + 1 for c in fill_col_names}
    input_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    # Recommendation colors
    rec_fills = {
        "Delete": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
        "Review for Deletion": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
        "Disable/Delete — Not Executing": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
    }

    rec_col = columns.index("recommendation") + 1

    for row_idx, rec in enumerate(cleanup, 3):
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=rec.get(col, ""))
            cell.font = BODY_FONT
            if col_idx in fill_col_indices:
                cell.fill = input_fill

        # Color the recommendation
        recommendation = rec.get("recommendation", "")
        rec_cell = ws.cell(row=row_idx, column=rec_col)
        for pattern, fill in rec_fills.items():
            if pattern in recommendation:
                rec_cell.fill = fill
                break

    col_widths = {"workspace": 25, "department": 20, "type": 10, "name": 55, "id": 38,
                  "staleness": 12, "days_since_update": 15, "has_lineage": 12,
                  "recommendation": 30, "assigned_to": 20, "action_taken": 15,
                  "action_date": 14, "notes": 40}
    for col_idx, col in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col, 15)

    ws.freeze_panes = "A3"


def _write_definition_interview_guide(ws) -> None:
    """Write a guide tab for the column definition interview process."""
    sections = [
        ("How to Use the Interview Tool", [
            ("Getting started", "Run: python3 interview.py --stats\nThis shows definition coverage by domain so you know where to focus."),
            ("Focus on your domain", "Run: python3 interview.py --domain \"Transactions\"\nReplace with your domain name. This filters to just your undefined columns."),
            ("Focus on one dataset", "Run: python3 interview.py --dataset \"Sites - Base Data Set\"\nGreat for tackling one important dataset at a time."),
            ("Offline review", "Run: python3 interview.py --export --domain \"RPA\"\nExports a CSV you can fill in offline, then import back:\npython3 interview.py --import-csv output/undefined_columns_rpa_20260411.csv"),
            ("Resume later", "Run: python3 interview.py --resume\nPicks up where you left off. Progress is saved automatically."),
        ]),
        ("Writing Good Definitions", [
            ("Start with what it IS", "Describe what the field represents, not where it comes from. 'Revenue generated from programmatic ad sales' not 'A revenue field'."),
            ("Include units/format", "If relevant: 'Revenue in USD', 'Duration in seconds', 'Date in YYYY-MM-DD format'."),
            ("Mention valid values", "For categorical fields: 'Site status (Active, Inactive, Temporarily Deactivated, Awaiting Installation, etc.)'"),
            ("No trailing periods", "Definitions should not end with a period. The tool enforces this automatically."),
            ("Source prefix", "If a column is specific to one source system, prefix with [Source]: '[Vistar] Daily fill rate percentage for programmatic ad slots'."),
            ("Skip with 's'", "If you don't know what a column is, skip it — don't guess. Someone else may know."),
            ("Mark N/A with 'n/a'", "For calculated/internal columns that don't need a business definition."),
        ]),
        ("Priority Order", [
            ("1. Your domain's Source datasets", "These are the foundational datasets everything else depends on. Get these right first."),
            ("2. Columns in 5+ datasets", "High-commonality columns affect the most users. The Column Dictionary tab (sorted by dataset_count) shows these."),
            ("3. Active datasets", "Don't spend time defining columns in Abandoned datasets that will be deleted."),
            ("4. Everything else", "Fill in as time permits."),
        ]),
        ("Domain Assignments", [
            ("Sites & Locations", "Data Engineering team"),
            ("Proof of Play", "Network Operations team"),
            ("Impressions", "Data & Analytics team"),
            ("Transactions", "Data & Analytics team"),
            ("Revenue & Monetization", "Finance team"),
            ("Programmatic / Programmatic Ops", "Programmatic team"),
            ("RPA", "Ad Operations team"),
            ("Traffic Instructions", "Ad Operations team"),
            ("Managed Services", "Ad Operations team"),
            ("Campaigns & Delivery", "Ad Operations team"),
            ("Salesforce / CRM", "Sales team"),
            ("Monitoring & Governance", "Data Engineering team"),
            ("Site Analytics", "Data & Analytics team"),
            ("Engineering", "Engineering team"),
        ]),
    ]

    row = 1
    for section_title, items in sections:
        ws.cell(row=row, column=1, value=section_title)
        ws.cell(row=row, column=1).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=2).fill = HEADER_FILL
        row += 1

        for term, desc in items:
            ws.cell(row=row, column=1, value=term).font = BOLD_FONT
            cell = ws.cell(row=row, column=2, value=desc)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            row += 1

        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 100
    ws.freeze_panes = "A2"


def _write_definition_issues(ws, issues_csv_path: str) -> None:
    """Write the Definition Issues tab from the validation CSV."""
    import os

    if not os.path.exists(issues_csv_path):
        # No issues file — write a placeholder
        ws.cell(row=1, column=1, value="No definition issues file found. Run validate_definitions.py first.")
        ws.cell(row=1, column=1).font = BOLD_FONT
        return

    rows = []
    with open(issues_csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        for row in reader:
            rows.append(row)

    if not rows:
        ws.cell(row=1, column=1, value="No definition issues found — all definitions passed validation.")
        ws.cell(row=1, column=1).font = BOLD_FONT
        return

    columns = ["issue_type", "column_name", "column_type", "current_definition", "suggestion"]

    # Instructions
    instruction_font = Font(name="Arial", size=10, bold=True, italic=True, color="1565C0")
    ws.cell(row=1, column=1, value=f"DEFINITION QUALITY ISSUES: {len(rows)} issues found across 5 checks. Review and resolve by priority (cross-type conflicts first, then too-short, then duplicates).")
    ws.cell(row=1, column=1).font = instruction_font

    # Headers
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Issue type colors
    issue_fills = {
        "conflicting_cross_type": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
        "too_short":              PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
        "generic_template":       PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
        "duplicate_definition":   PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
        "prefix_mismatch":        PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid"),
    }

    for row_idx, rec in enumerate(rows, 3):
        issue_type = rec.get("issue_type", "")
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=rec.get(col, ""))
            cell.font = BODY_FONT
            if col == "current_definition" or col == "suggestion":
                cell.alignment = WRAP
            # Color the issue_type cell
            if col_idx == 1:
                for pattern, fill in issue_fills.items():
                    if pattern in issue_type:
                        cell.fill = fill
                        break

    col_widths = {"issue_type": 25, "column_name": 35, "column_type": 12, "current_definition": 70, "suggestion": 60}
    for col_idx, col in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col, 15)

    ws.freeze_panes = "A3"


def add_runbook_tabs(workbook_path: str) -> None:
    """Add runbook and interview tabs to the existing workspace planning workbook."""
    cache = load_cache()
    if cache is None:
        raise RuntimeError("No cache found.")

    datasets = cache["datasets"]
    lineage = cache["lineage"]
    dataflows = cache["dataflows"]

    wb = load_workbook(workbook_path)

    # Remove existing tabs if re-running
    for name in ["Runbook", "Unclassified Review", "Owner Action Map", "Definition Guide", "Definition Issues"]:
        if name in wb.sheetnames:
            del wb[name]

    # Tab: Runbook
    ws_runbook = wb.create_sheet("Runbook")
    _write_runbook(ws_runbook)

    # Tab: Unclassified Review
    ws_unclass = wb.create_sheet("Unclassified Review")
    _write_unclassified_review(ws_unclass, datasets, lineage)

    # Tab: Owner Action Map
    ws_owner = wb.create_sheet("Owner Action Map")
    _write_owner_action_map(ws_owner, datasets, lineage, dataflows)

    # Tab: Definition Guide
    ws_guide = wb.create_sheet("Definition Guide")
    _write_definition_interview_guide(ws_guide)

    # Tab: Definition Issues
    issues_csv = str(Path(__file__).parent / "output" / "definition_issues_20260411.csv")
    ws_issues = wb.create_sheet("Definition Issues")
    _write_definition_issues(ws_issues, issues_csv)

    wb.save(workbook_path)
    logger.info("Added runbook tabs to %s", workbook_path)


if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.WARNING)

    workbook_path = "output/domo_workspace_plan_20260411.xlsx"
    if len(sys.argv) > 1:
        workbook_path = sys.argv[1]

    print(f"Adding runbook and interview tabs to {workbook_path}...", end=" ", flush=True)
    add_runbook_tabs(workbook_path)
    print("✓")
