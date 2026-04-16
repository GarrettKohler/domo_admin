#!/usr/bin/env python3
"""Apply approved dataset renames via the Domo API.

Reads returned cleanup spreadsheets (or the master rename CSVs) and pushes
approved name changes to Domo. Supports dry-run mode for validation.

Usage:
    # Dry run — show what would be renamed, no changes made
    python3 apply_renames.py --dry-run

    # Execute renames from returned cleanup spreadsheets
    python3 apply_renames.py --source spreadsheets

    # Execute renames from master CSV (conservative)
    python3 apply_renames.py --source csv --mode conservative

    # Execute renames from master CSV (aggressive)
    python3 apply_renames.py --source csv --mode aggressive
"""

import argparse
import csv
import json
import logging
import os
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from domo_client import DomoClient

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

OUTPUT_DIR = Path(__file__).resolve().parent / "output"
ROLLOUT_DIR = OUTPUT_DIR / "owner_rollouts"
LOG_DIR = OUTPUT_DIR / "automation_logs"


def load_renames_from_spreadsheets() -> list[dict]:
    """Load approved renames from returned owner cleanup spreadsheets.

    Looks for spreadsheets in the rollout directory and reads rows where:
    - Decision column is "Keep" (or blank — rename either way if approved)
    - Proposed Name or Restructured Name differs from current name
    - Owner has indicated a rename preference (or accepted default)
    """
    renames = []

    if not ROLLOUT_DIR.exists():
        logger.error("Rollout directory not found: %s", ROLLOUT_DIR)
        return renames

    for xlsx_path in sorted(ROLLOUT_DIR.glob("cleanup_review_*.xlsx")):
        logger.info("Reading %s", xlsx_path.name)
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)

        for sheet_name in ("Datasets", "Dataflows"):
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            headers = [str(h).strip() if h else "" for h in rows[0]]

            # Find column indices
            col_map = {h.lower(): i for i, h in enumerate(headers)}
            id_col = col_map.get("id") or col_map.get("dataset id") or col_map.get("dataflow id")
            name_col = col_map.get("name") or col_map.get("dataset name") or col_map.get("dataflow name")
            proposed_col = col_map.get("proposed name")
            restructured_col = col_map.get("restructured name")
            decision_col = col_map.get("decision")

            if id_col is None or name_col is None:
                logger.warning("  Skipping %s — missing ID or Name column", sheet_name)
                continue

            for row in rows[1:]:
                if not row or not row[id_col]:
                    continue

                item_id = str(row[id_col]).strip()
                current_name = str(row[name_col]).strip() if row[name_col] else ""
                decision = str(row[decision_col]).strip().lower() if decision_col is not None and row[decision_col] else ""

                # Skip items marked for removal
                if decision == "remove":
                    continue

                # Only rename datasets (dataflows can't be renamed via API)
                item_type = "dataset" if sheet_name == "Datasets" else "dataflow"

                # Determine the target name — prefer restructured if populated
                new_name = None
                if restructured_col is not None and row[restructured_col]:
                    new_name = str(row[restructured_col]).strip()
                elif proposed_col is not None and row[proposed_col]:
                    new_name = str(row[proposed_col]).strip()

                if new_name and new_name != current_name:
                    renames.append({
                        "id": item_id,
                        "current_name": current_name,
                        "new_name": new_name,
                        "type": item_type,
                        "source": xlsx_path.name,
                    })

        wb.close()

    return renames


def load_renames_from_csv(mode: str = "conservative") -> list[dict]:
    """Load renames from the master rename CSV files.

    mode: 'conservative' or 'aggressive'
    """
    renames = []

    if mode == "aggressive":
        files = [
            ("dataset", OUTPUT_DIR.parent / "dataset_aggressive_renames.csv"),
            ("dataflow", OUTPUT_DIR.parent / "dataflow_aggressive_renames.csv"),
        ]
    else:
        files = [
            ("dataset", OUTPUT_DIR.parent / "dataset_renames.csv"),
            ("dataflow", OUTPUT_DIR.parent / "dataflow_renames.csv"),
        ]

    for item_type, csv_path in files:
        if not csv_path.exists():
            logger.warning("Rename CSV not found: %s", csv_path)
            continue

        with open(csv_path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                item_id = row.get("dataset_id") or row.get("dataflow_id") or row.get("id", "")
                current = row.get("current_name") or row.get("original_name", "")
                proposed = row.get("proposed_name") or row.get("new_name", "")

                if item_id and proposed and proposed != current:
                    renames.append({
                        "id": item_id,
                        "current_name": current,
                        "new_name": proposed,
                        "type": item_type,
                        "source": csv_path.name,
                    })

    return renames


def apply_renames(renames: list[dict], dry_run: bool = True) -> dict:
    """Apply renames via the Domo API.

    Returns summary stats dict.
    """
    # Filter to datasets only — dataflow renames require UI
    dataset_renames = [r for r in renames if r["type"] == "dataset"]
    dataflow_renames = [r for r in renames if r["type"] == "dataflow"]

    if dataflow_renames:
        logger.warning(
            "Skipping %d dataflow renames — dataflows cannot be renamed via API",
            len(dataflow_renames),
        )

    stats = {
        "total": len(dataset_renames),
        "success": 0,
        "failed": 0,
        "skipped_dataflows": len(dataflow_renames),
        "dry_run": dry_run,
    }

    if not dataset_renames:
        logger.info("No dataset renames to apply.")
        return stats

    # Create log directory
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = LOG_DIR / f"renames_{timestamp}.csv"

    if dry_run:
        logger.info("\n" + "=" * 70)
        logger.info("DRY RUN — No changes will be made")
        logger.info("=" * 70)
        logger.info("Would rename %d datasets:\n", len(dataset_renames))

        with open(log_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["dataset_id", "current_name", "new_name", "status"])
            for r in dataset_renames:
                logger.info("  %s", r["id"])
                logger.info("    FROM: %s", r["current_name"])
                logger.info("    TO:   %s", r["new_name"])
                writer.writerow([r["id"], r["current_name"], r["new_name"], "dry_run"])

        logger.info("\nDry run log saved to: %s", log_path)
        return stats

    # Live execution
    client = DomoClient(
        client_id=os.environ["DOMO_CLIENT_ID"],
        client_secret=os.environ["DOMO_CLIENT_SECRET"],
    )
    client.authenticate()
    logger.info("Authenticated with Domo API")

    results = []
    for i, r in enumerate(dataset_renames, 1):
        logger.info("[%d/%d] Renaming %s → %s", i, len(dataset_renames), r["current_name"][:50], r["new_name"][:50])

        result = client.update_dataset(r["id"], name=r["new_name"])
        if result:
            stats["success"] += 1
            results.append([r["id"], r["current_name"], r["new_name"], "success"])
            logger.info("  ✓ Success")
        else:
            stats["failed"] += 1
            results.append([r["id"], r["current_name"], r["new_name"], "failed"])
            logger.error("  ✗ Failed")

    # Write results log
    with open(log_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["dataset_id", "current_name", "new_name", "status"])
        writer.writerows(results)

    logger.info("\nResults log saved to: %s", log_path)
    return stats


def main():
    parser = argparse.ArgumentParser(description="Apply approved dataset renames to Domo")
    parser.add_argument("--dry-run", action="store_true", default=True,
                        help="Preview changes without applying (default: True)")
    parser.add_argument("--execute", action="store_true",
                        help="Actually apply the renames (overrides --dry-run)")
    parser.add_argument("--source", choices=["spreadsheets", "csv"], default="spreadsheets",
                        help="Where to read renames from (default: spreadsheets)")
    parser.add_argument("--mode", choices=["conservative", "aggressive"], default="conservative",
                        help="Rename mode when using --source csv (default: conservative)")
    args = parser.parse_args()

    dry_run = not args.execute

    if args.source == "spreadsheets":
        renames = load_renames_from_spreadsheets()
    else:
        renames = load_renames_from_csv(args.mode)

    logger.info("Loaded %d rename candidates", len(renames))

    if not renames:
        logger.info("Nothing to rename. Exiting.")
        return

    stats = apply_renames(renames, dry_run=dry_run)

    print(f"\n{'=' * 60}")
    print("RENAME SUMMARY")
    print(f"{'=' * 60}")
    print(f"  Mode:              {'DRY RUN' if stats['dry_run'] else 'LIVE'}")
    print(f"  Dataset renames:   {stats['total']}")
    if not dry_run:
        print(f"  Successful:        {stats['success']}")
        print(f"  Failed:            {stats['failed']}")
    print(f"  Dataflows skipped: {stats['skipped_dataflows']} (requires UI)")


if __name__ == "__main__":
    main()
